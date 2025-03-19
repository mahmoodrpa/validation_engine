from django.shortcuts import render, HttpResponse, redirect
from django.contrib import messages
import pandas as pd
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.template import loader
from dateutil import parser
from pathlib import Path
from django.core.paginator import Paginator
from django.shortcuts import render
from .forms import *
from .models import *
import pandas as pd
from datetime import datetime
from django.db.models import Count, OuterRef, Subquery
from django.db.models import Sum, F, Value
from django.db.models.functions import Concat
from django.db import transaction
from django.db.models import Q
from django.core.exceptions import ValidationError
from decimal import Decimal
from django.views.decorators.http import condition
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
from django.utils import timezone
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponseServerError
import time
from openpyxl import Workbook
from main import models as mm
from django.db import IntegrityError
from django.db.models import Sum
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import io
import base64
from decimal import Decimal

# Create your views here.

@login_required(login_url="/login/")
def amazon_remit_upload(request, excel_file, context):
    try:    
        print("Entered try block")
        if excel_file:
            # Load the Excel data into a pandas DataFrame
            # df = pd.read_excel(excel_file)
            df = pd.read_csv(excel_file)
            df['unique_key'] = df['Payment Number'].astype(str) + df['Invoice Number'].astype(str)
            df_unique = df.drop_duplicates(subset=['unique_key'])
            df_unique = df_unique.drop(columns=['unique_key'])
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df_unique.shape[0]
            instances = []
            # Iterate over each row in the DataFrame and validate the data
            for _, row in df_unique.iterrows():
                payment_number = row['Payment Number']
                invoice_number = row['Invoice Number']
                invoice_date= row['Invoice Date']
                if pd.isnull(invoice_date):
                    # Handle missing value
                    invoice_date = None
                else:
                    invoice_date = parse_date(row['Invoice Date'])
                description = row['Description']
                invoice_amount = row['Invoice Amount']
                if pd.isnull(invoice_amount):
                    invoice_amount=0
                else:
                    invoice_amount=convert_to_decimal_V2(row['Invoice Amount'])
                invoice_currency = row['Invoice Currency']
                withholding_amount = row['Withholding Amount']
                if pd.isnull(withholding_amount):
                    withholding_amount=0
                else:
                    withholding_amount=convert_to_decimal_V2(row['Withholding Amount'])
                terms_discount = row['Terms Discount Taken']
                if pd.isnull(terms_discount):
                    terms_discount=0
                else:
                    terms_discount=convert_to_decimal_V2(row['Terms Discount Taken'])
                deduction_amount = row['Amount Paid']
                if pd.isnull(deduction_amount):
                    deduction_amount=0
                else:
                    deduction_amount=convert_to_decimal_V2(row['Amount Paid'])

                remaining_amount = row['Remaining Amount']
                if pd.isnull(remaining_amount):
                    remaining_amount=0
                else:
                    remaining_amount=convert_to_decimal_V2(row['Remaining Amount'])
                payment_date = row['Payment Date']
                if pd.isnull(payment_date):
                    payment_date = None
                else:
                    payment_date = parse_date(row['Payment Date'])
                payment_amount = row['Payment Amount']
                if pd.isnull(payment_amount):
                    payment_amount=0
                else:
                    payment_amount=convert_to_decimal_V2(row['Payment Amount'])
                deduction_reason = row['Reason Code']
                sub_reason = row['Sub Reason Code']

                # Check if the record already exists in the deductionsdata model
                if amazonremittance.objects.filter(invoice_number=invoice_number,deduction_amount=deduction_amount).exists():
                    # Skip duplicates
                    continue
                instance = amazonremittance(
                    payment_number = payment_number,
                    invoice_number = invoice_number,
                    invoice_date= invoice_date,
                    description = description,
                    invoice_amount = invoice_amount,
                    invoice_currency = invoice_currency,
                    withholding_amount = withholding_amount,
                    terms_discount = terms_discount,
                    deduction_amount = deduction_amount,
                    remaining_amount = remaining_amount,
                    payment_date = payment_date,
                    payment_amount = payment_amount,
                    deduction_reason = deduction_reason,
                    sub_reason = sub_reason
                


                )
                # Amazonpodata.save()
                instances.append(instance)

            # Bulk insert the instances
            amazonremittance.objects.bulk_create(instances)
            end_time = timezone.now()
            total_time = end_time - start_time

            # Update status to 'completed'
            current_status = 'completed'
            request.session['upload_status'] = 'success'
            
            #update upload status
            uploaded_csv = mm.UploadedCSV.objects.create(
            user=request.user,
            file=excel_file,
            num_rows=df_unique.shape[0],
            start_time=start_time,
            end_time=end_time,
            total_time=total_time,
            current_status=current_status
            )
            return redirect('backup_upload')
        else:
            # Handle the case when excel_file is not present
            # You can show an error message or take appropriate action
            print("File not present")
            context['error'] = 'No file was uploaded.'
    except pd.errors.EmptyDataError:
        print("EmptyDataError caught")
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        print(f'ParserError caught: {e}')
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        print(f'ValidationError caught: {e}')
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        # Handle other unexpected exceptions
        print(f'Unexpected error caught: {e}')
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)    
    # pass

def convert_to_decimal(value):
    if isinstance(value, str):
        try:
            return float(value.replace(',', ''))
        except ValueError:
            return None
    elif isinstance(value, float):
        return value
    else:
        return None

def convert_to_decimal_V2(value):    
    if isinstance(value, str):
        try:
            # Remove dollar signs and commas, then convert to float
            cleaned_value = value.replace('$', '').replace(',', '')
            result = float(cleaned_value)
            return result
        except ValueError:
            return None
    elif isinstance(value, (int, float)):
        return float(value)
    else:
        return None

def round_rebate(rebate_value, decimal_places=2):
    if rebate_value is not None:
        return round(rebate_value, decimal_places)
    return None

def parse_date(date_str):
    if pd.isnull(date_str):
        return None
    if not isinstance(date_str, str):
        date_str = str(date_str).strip()
    else:
        date_str = date_str.strip()

    date_formats = [
        '%m/%d/%Y', '%m-%d-%Y', '%d-%m-%Y', '%d-%b-%y', '%d-%b-%Y',
        '%m/%d/%y', '%Y/%m/%d', '%Y-%m-%d', '%Y-%m-%dT%H:%M:%S.%fZ',
        '%Y-%m-%d %H:%M:%S', '%d-%m-%y', '%m-%d-%y'
    ]

    for fmt in date_formats:
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue

    print(f"Error parsing date: {date_str}")
    return None

@login_required(login_url="/login/")
def amazon_promo_backup_upload(request, excel_file, context):
    try:    
        print("Entered try block")
        if excel_file:
            
            encodings_to_try = ['utf-8', 'latin-1', 'iso-8859-1']
            df = None

            for encoding in encodings_to_try:
                try:
                    df = pd.read_csv(excel_file, encoding=encoding, dtype={'UPC': str, 'EAN': str, 'Invoice': str, 'Unique_Key':str})
                    df['unique_key'] = df['Purchase Order'].astype(str) + df['Asin'].astype(str) + df['Receive Date'].astype(str)
                    df_unique = df.drop_duplicates(subset=['unique_key'])
                    df_unique = df_unique.drop(columns=['unique_key'])
                    break  # Stop trying other encodings if one works
                except UnicodeDecodeError:
                    pass

            if df is None:
                # Handle the case when no encoding works
                context['error'] = 'Unable to decode the CSV file using available encodings.'
                return render(request, 'backup_upload.html')

            # Record start time
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df_unique.shape[0]
            
            instances = []
            # Iterate over each row in the DataFrame and validate the data
            for _, row in df.iterrows():
                Deduction_ref=row['Invoice ID']
                Receive_Date=row['Receive Date']
                if pd.isnull(Receive_Date):
                    # Handle missing value
                    Receive_Date = None
                else:
                    Receive_Date = parse_date(row['Receive Date'])
                
                Return_Date=row['Return Date']            
                if pd.isnull(Return_Date):
                    # Handle missing datetime value
                    Return_Date = None
                else: 
                    Return_Date = parse_date(row['Return Date'])
                
                Invoice_Day=row['Invoice Day'] 
                if pd.isnull(Invoice_Day):
                    # Handle missing value
                    Invoice_Day = None
                else:
                    Invoice_Day = parse_date(row['Invoice Day'])
                
                Transaction_Type=row['Transaction Type']
                
                Quantity=convert_to_decimal(row['Quantity'])
                if pd.isnull(Quantity):
                    # Handle missing value
                    Quantity = None
                else:
                    Quantity=convert_to_decimal(row['Quantity'])  
                
                Net_Receipts=convert_to_decimal(row['Net Receipts'])
                if pd.isnull(Net_Receipts):
                    # Handle missing value
                    Net_Receipts = None
                else:
                    Net_Receipts=convert_to_decimal(row['Net Receipts'])  

                Net_Receipts_Currency=row['Net Receipts Currency']

                List_Price=convert_to_decimal(row['List Price'])
                if pd.isnull(List_Price):
                    # Handle missing value
                    List_Price = None
                else:
                    List_Price=convert_to_decimal(row['List Price'])  

                List_Price_Currency=row['List Price Currency']
                
                Rebate_In_Agreement_Currency=convert_to_decimal(row['Rebate In Agreement Currency'])
                if pd.isnull(Rebate_In_Agreement_Currency):
                    # Handle missing value
                    Rebate_In_Agreement_Currency = None
                else:
                    Rebate_In_Agreement_Currency=convert_to_decimal(row['Rebate In Agreement Currency'])

                Agreement_Currency=row['Agreement Currency']
                
                Rebate_In_Purchase_Order_Currency=convert_to_decimal(row['Rebate In Purchase Order Currency'])
                if pd.isnull(Rebate_In_Purchase_Order_Currency):
                    # Handle missing value
                    Rebate_In_Purchase_Order_Currency = None
                else:
                    Rebate_In_Purchase_Order_Currency=convert_to_decimal(row['Rebate In Purchase Order Currency'])
                
                Purchase_Order_Currency=row['Purchase Order Currency']
                if pd.isnull(Purchase_Order_Currency):
                    # Handle missing value
                    Purchase_Order_Currency = None
                else:
                    Purchase_Order_Currency=row['Purchase Order Currency']

                Purchase_Order=row['Purchase Order']
                Promo_Asin=row['Asin']
                UPC=row['UPC']
                EAN=str(row['EAN'])
                Manufacturer=row['Manufacturer']
                Distributor=row['Distributor']
                Product_Group=row['Product Group']
                Category_=row['Category']
                Subcategory_=row['Subcategory']
                Title=row['Title']
                Product_Description=row['Product Description']
                Binding=row['Binding']
                Cost_Currency=row['Cost Currency']

                Old_Cost=row['Old Cost']
                if pd.isnull(Old_Cost):
                    # Handle missing value
                    Old_Cost = None
                else:
                    Old_Cost=row['Old Cost']

                New_Cost=row['New Cost']
                if pd.isnull(New_Cost):
                    # Handle missing value
                    New_Cost = None
                else:
                    New_Cost=row['New Cost']

                Price_Protection_Agreement=row['Price Protection Agreement']

                Price_Protection_Day=row['Price Protection Day']
                if pd.isnull(Price_Protection_Day):
                    # Handle missing value
                    Price_Protection_Day = None
                else:
                    Price_Protection_Day=row['Price Protection Day']

                Cost_Variance=convert_to_decimal(row['Cost Variance'])
                if pd.isnull(Cost_Variance):
                    # Handle missing value
                    Cost_Variance = None
                else:
                    Cost_Variance=convert_to_decimal(row['Cost Variance'])

                Invoice=str(row['Invoice'])

                
                Deduction_ref = row['Invoice ID']
                # Check if the record already exists in the deductionsdata model
                if amazonpromo.objects.filter(Purchase_Order=Purchase_Order, Promo_Asin=Promo_Asin,Rebate_In_Agreement_Currency=round_rebate(Rebate_In_Agreement_Currency) ).exists():
                    # Skip duplicates
                    # print(f"Duplicate record found - Purchase_Order: {Purchase_Order}, Promo_Asin: {Promo_Asin}, Rebate_In_Agreement_Currency: {Rebate_In_Agreement_Currency}")    
                    continue
                instance = amazonpromo(
                    Deduction_ref=Deduction_ref,
                    Receive_Date=Receive_Date,
                    Return_Date=Return_Date,
                    Invoice_Day=Invoice_Day,
                    Transaction_Type=Transaction_Type,
                    Quantity=Quantity,
                    Net_Receipts=Net_Receipts,
                    Net_Receipts_Currency=Net_Receipts_Currency,
                    List_Price=List_Price,
                    List_Price_Currency=List_Price_Currency,
                    Rebate_In_Agreement_Currency=Rebate_In_Agreement_Currency,
                    Agreement_Currency=Agreement_Currency,
                    Rebate_In_Purchase_Order_Currency=Rebate_In_Purchase_Order_Currency,
                    Purchase_Order_Currency=Purchase_Order_Currency,
                    Purchase_Order=Purchase_Order,
                    Promo_Asin=Promo_Asin,
                    UPC=UPC,
                    EAN=EAN,
                    Manufacturer=Manufacturer,
                    Distributor=Distributor,
                    Product_Group=Product_Group,
                    Category=Category_,
                    Subcategory=Subcategory_,
                    Title=Title,
                    Product_Description=Product_Description,
                    Binding=Binding,
                    Cost_Currency=Cost_Currency,
                    Old_Cost=Old_Cost,
                    New_Cost=New_Cost,
                    Price_Protection_Agreement=Price_Protection_Agreement,
                    Price_Protection_Day=Price_Protection_Day,
                    Cost_Variance=Cost_Variance,
                    Invoice=Invoice

                )
                # print("New Instance Data:")
                # print(f"Purchase_Order: {instance.Purchase_Order}")
                # print(f"Promo_Asin: {instance.Promo_Asin}")
                # print(f"Rebate_In_Agreement_Currency:{instance.Rebate_In_Agreement_Currency}")
                instances.append(instance)
                # print("INSTANCE",instance,"instance")
            # Bulk insert the instances
            amazonpromo.objects.bulk_create(instances)
            # Log the non-duplicate entries
            # print(f"Non-duplicate record - Purchase_Order: {Purchase_Order}, Promo_Asin: {Promo_Asin}, Rebate_In_Agreement_Currency: {Rebate_In_Agreement_Currency}")
            end_time = timezone.now()
            total_time = end_time - start_time

            # Update status to 'completed'
            current_status = 'completed'
            request.session['upload_status'] = 'success'

            uploaded_csv = mm.UploadedCSV.objects.create(
            user=request.user,
            file=excel_file,
            num_rows=df_unique.shape[0],
            start_time=start_time,
            end_time=end_time,
            total_time=total_time,
            current_status=current_status
            )

            return redirect('backup_upload')
        else:
            # Handle the case when excel_file is not present
            # You can show an error message or take appropriate action
            print("File not present")
            context['error'] = 'No file was uploaded.'
    except pd.errors.EmptyDataError:
        print("EmptyDataError caught")
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        print(f'ParserError caught: {e}')
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        print(f'ValidationError caught: {e}')
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        # Handle other unexpected exceptions
        print(f'Unexpected error caught: {e}')
        context['error'] = f'An unexpected error occurred: {e}'        
    return render(request, 'backup_upload.html')

@login_required(login_url="/login/")
def amazon_po_upload(request, excel_file, context):
    try:
        print("Entered try block")
        if excel_file:
            # Load the Excel data into a pandas DataFrame
            # df = pd.read_excel(excel_file)
            df = pd.read_csv(excel_file)
            df['unique_key'] = df['PO'].astype(str) + df['ASIN'].astype(str)
            df_unique = df.drop_duplicates(subset=['unique_key'])
            df_unique = df_unique.drop(columns=['unique_key'])
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df_unique.shape[0]
            instances = []
            # Iterate over each row in the DataFrame and validate the data
            for _, row in df_unique.iterrows():
                PO=row['PO']
                PO_ASIN=row['ASIN']
                External_ID=row['External ID']
                External_Id_Type=row['External Id Type']
                Model_Number=row['Model Number']
                Title=row['Title']
                Availability=row['Availability']
                Backordered=row['Backordered']
                Window_Type=row['Window Type']
                Window_Start=row['Window Start']
                Window_End=row['Window End']
                Expected_Date=row['Expected Date']
                Case_Size=row['Case Size']
                Quantity_Requested=row['Quantity Requested']
                Accepted_Quantity=row['Accepted Quantity']
                Quantity_received=row['Quantity received']
                Quantity_Outstanding=row['Quantity Outstanding']
                Case_Cost=row['Case Cost']
                Total_Cost=row['Total Cost']
                
                if pd.isnull(Case_Size):
                    # Handle missing value
                    Case_Size = None
                else:
                    Case_Size=row['Case Size']
                
                if pd.isnull(Quantity_Requested):
                    # Handle missing value
                    Quantity_Requested = None
                else:
                    Quantity_Requested=row['Quantity Requested']
                
                if pd.isnull(Accepted_Quantity):
                    # Handle missing value
                    Accepted_Quantity = None
                else:
                    Accepted_Quantity=row['Accepted Quantity']

                if pd.isnull(Quantity_received):
                    # Handle missing value
                    Quantity_received = None
                else:
                    Quantity_received=row['Quantity received']

                if pd.isnull(Quantity_Outstanding):
                    # Handle missing value
                    Quantity_Outstanding = None
                else:
                    Quantity_Outstanding=row['Quantity Outstanding']
                
                if pd.isnull(Case_Cost):
                    # Handle missing value
                    Case_Cost = None
                else:
                    Case_Cost=row['Case Cost']

                if pd.isnull(Total_Cost):
                    # Handle missing value
                    Total_Cost = None
                else:
                    Total_Cost=row['Total Cost']

                if pd.isnull(Window_Start):
                    # Handle missing value
                    Window_Start = None
                else:
                    # Window_Start = datetime.strptime(Window_Start, '%m/%d/%Y').date()
                    Window_Start = parse_date(row['Window Start'])
                
                if pd.isnull(Window_End):
                    # Handle missing value
                    Window_End = None
                else:
                    # Window_End = datetime.strptime(Window_End, '%m/%d/%Y').date()
                    Window_End = parse_date(row['Window End'])
                
                if pd.isnull(Expected_Date):
                    # Handle missing value
                    Expected_Date = None
                else:
                    # Expected_Date = datetime.strptime(Expected_Date, '%m/%d/%Y').date()
                    Expected_Date = parse_date(row['Expected Date'])


                # Check if the record already exists in the deductionsdata model
                if amazonpodata.objects.filter(PO=PO, PO_ASIN=PO_ASIN).exists():
                    # Skip duplicates
                    continue
                instance = amazonpodata(
                    PO=PO,
                    PO_ASIN=PO_ASIN,
                    External_ID=External_ID,
                    # External_Id_Type=External_Id_Type,
                    Model_Number=Model_Number,
                    Title=Title,
                    Availability=Availability,
                    Backordered=Backordered,
                    Window_Type=Window_Type,
                    Window_Start=Window_Start,
                    Window_End=Window_End,
                    Expected_Date=Expected_Date,
                    Case_Size=Case_Size,
                    Quantity_Requested=Quantity_Requested,
                    Accepted_Quantity=Accepted_Quantity,
                    Quantity_received=Quantity_received,
                    Quantity_Outstanding=Quantity_Outstanding,
                    Case_Cost=Case_Cost,
                    Total_Cost=Total_Cost,
                    )
                # Amazonpodata.save()
                instances.append(instance)

            # Bulk insert the instances
            amazonpodata.objects.bulk_create(instances)
            end_time = timezone.now()
            total_time = end_time - start_time

            # Update status to 'completed'
            current_status = 'completed'
            request.session['upload_status'] = 'success'
            
            #update upload status
            uploaded_csv = mm.UploadedCSV.objects.create(
            user=request.user,
            file=excel_file,
            num_rows=df_unique.shape[0],
            start_time=start_time,
            end_time=end_time,
            total_time=total_time,
            current_status=current_status
            )
            return redirect('backup_upload')
        else:
            # Handle the case when excel_file is not present
            # You can show an error message or take appropriate action
            context['error'] = 'No Excel file was uploaded.'
        # pass
    except pd.errors.EmptyDataError:
        print("EmptyDataError caught")
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        print(f'ParserError caught: {e}')
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        print(f'ValidationError caught: {e}')
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        # Handle other unexpected exceptions
        print(f'Unexpected error caught: {e}')
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)        

@login_required(login_url="/login/")
def amazon_agreement_upload(request, excel_file, context):
    try:
        print("Entered try block")
        if excel_file:
            df = pd.read_csv(excel_file)
            df['unique_key'] = df['INVOICE NUMBER'].astype(str) + df['AGREEMENT#'].astype(str)
            df_unique = df.drop_duplicates(subset=['unique_key'])
            df_unique = df_unique.drop(columns=['unique_key'])
            start_time = timezone.now()
            # Set initial status
            current_status = 'started'
            total_rows = df_unique.shape[0]
            instances = []
            # Iterate over each row in the DataFrame and validate the data
            for _, row in df_unique.iterrows():
                INVOICE_NUMBER=row['INVOICE NUMBER']
                INVOICE_DATE=row['INVOICE DATE']
                if pd.isnull(INVOICE_DATE):
                    # Handle missing datetime value
                    INVOICE_DATE = None
                else:
                    # INVOICE_DATE=datetime.strptime(INVOICE_DATE, '%m/%d/%Y').date()
                    INVOICE_DATE = parse_date(row['INVOICE DATE'])
                PAYMENT_TERM=row['PAYMENT TERM']
                DUE_DATE=row['DUE DATE']
                if pd.isnull(DUE_DATE):
                    # Handle missing datetime value
                    DUE_DATE = None
                else:
                    # DUE_DATE=datetime.strptime(DUE_DATE, '%m/%d/%Y').date()
                    DUE_DATE = parse_date(row['INVOICE DATE'])
                BUYER=row['BUYER']
                PUB_CODE=row['PUB CODE']
                AGREEMENT_num=row['AGREEMENT#']
                PAYMENT_METHOD=row['PAYMENT METHOD']
                TRANSACTION_TYPE=row['TRANSACTION TYPE']
                PRODUCT_LINE=row['PRODUCT LINE']
                PO_NUMBER=row['PO NUMBER']
                if pd.isnull(PO_NUMBER):
                    # Handle missing value
                    PO_NUMBER = None
                else:
                    PO_NUMBER=convert_to_decimal(row['PO NUMBER'])
                LINE_NUMBER=row['LINE NUMBER']
                QUANTITY_ORDERED=row['QUANTITY ORDERED']
                QUANTITY_CREDITED=row['QUANTITY CREDITED']
                if pd.isnull(QUANTITY_CREDITED):
                    # Handle missing value
                    QUANTITY_CREDITED = None
                else:
                    QUANTITY_CREDITED=convert_to_decimal(row['QUANTITY CREDITED'])
                QUANTITY_INVOICED=row['QUANTITY INVOICED']
                if pd.isnull(QUANTITY_INVOICED):
                    # Handle missing value
                    QUANTITY_INVOICED = None
                else:
                    QUANTITY_INVOICED=convert_to_decimal(row['QUANTITY INVOICED'])
                UNIT_STANDARD_PRICE=row['UNIT STANDARD PRICE']
                if pd.isnull(UNIT_STANDARD_PRICE):
                    # Handle missing value
                    UNIT_STANDARD_PRICE = None
                else:
                    UNIT_STANDARD_PRICE=convert_to_decimal(row['UNIT STANDARD PRICE'])
                UNIT_SELLING_PRICE=row['UNIT SELLING PRICE']
                REVENUE_AMOUNT=row['REVENUE AMOUNT']
                SALES_ORDER=row['SALES ORDER']
                if pd.isnull(SALES_ORDER):
                    # Handle missing value
                    SALES_ORDER = None
                else:
                    SALES_ORDER=convert_to_decimal(row['SALES ORDER'])
                SALES_ORDER_LINE=row['SALES ORDER LINE']
                if pd.isnull(SALES_ORDER_LINE):
                    # Handle missing value
                    SALES_ORDER_LINE = None
                else:
                    SALES_ORDER_LINE=convert_to_decimal(row['SALES ORDER LINE'])
                DESCRIPTION=row['DESCRIPTION']
                REBATE_prct=row['REBATE%']
                # REBATE_prct=Decimal(REBATE_prct.strip('%'))
                REBATE_prct=REBATE_prct/100
                AGREEMENT_END=row['AGREEMENT END']
                if pd.isnull(AGREEMENT_END):
                    # Handle missing datetime value
                    AGREEMENT_END = None
                else:
                    # AGREEMENT_END=datetime.strptime(AGREEMENT_END, '%m/%d/%Y').date() 
                    AGREEMENT_END = parse_date(row['AGREEMENT END']) 
                
                INVOICE_NUMBER = str(INVOICE_NUMBER)[:14]

                # INVOICE_NUMBER = amazonremit.objects.get(Customer_Ref=INVOICE_NUMBER)
                # Check if the record already exists in the deductionsdata model
                if amazonagreement.objects.filter(INVOICE_NUMBER=INVOICE_NUMBER).exists():
                    # Skip duplicates
                    continue
                instance = amazonagreement(
                    INVOICE_NUMBER=INVOICE_NUMBER,
                    INVOICE_DATE=INVOICE_DATE,
                    PAYMENT_TERM=PAYMENT_TERM,
                    DUE_DATE=DUE_DATE,
                    BUYER=BUYER,
                    PUB_CODE=PUB_CODE,
                    AGREEMENT_num=AGREEMENT_num,
                    PAYMENT_METHOD=PAYMENT_METHOD,
                    TRANSACTION_TYPE=TRANSACTION_TYPE,
                    PRODUCT_LINE=PRODUCT_LINE,
                    PO_NUMBER=PO_NUMBER,
                    LINE_NUMBER=LINE_NUMBER,
                    QUANTITY_ORDERED=QUANTITY_ORDERED,
                    QUANTITY_CREDITED=QUANTITY_CREDITED,
                    QUANTITY_INVOICED=QUANTITY_INVOICED,
                    UNIT_STANDARD_PRICE=UNIT_STANDARD_PRICE,
                    UNIT_SELLING_PRICE=UNIT_SELLING_PRICE,
                    REVENUE_AMOUNT=REVENUE_AMOUNT,
                    SALES_ORDER=SALES_ORDER,
                    SALES_ORDER_LINE=SALES_ORDER_LINE,
                    DESCRIPTION=DESCRIPTION,
                    REBATE_prct=REBATE_prct,
                    AGREEMENT_END=AGREEMENT_END,
                )
                instances.append(instance)
                    # Bulk insert the instances
            amazonagreement.objects.bulk_create(instances)
            end_time = timezone.now()
            total_time = end_time - start_time

            # Update status to 'completed'
            current_status = 'completed'
            request.session['upload_status'] = 'success'
            
            #update upload status
            uploaded_csv = mm.UploadedCSV.objects.create(
            user=request.user,
            file=excel_file,
            num_rows=df_unique.shape[0],
            start_time=start_time,
            end_time=end_time,
            total_time=total_time,
            current_status=current_status
            )
            return redirect('backup_upload')
        else:
            # Handle the case when excel_file is not present
            # You can show an error message or take appropriate action
            print("File not present")
            context['error'] = 'No file was uploaded.'
    
            # pass
    except pd.errors.EmptyDataError:
        print("EmptyDataError caught")
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        print(f'ParserError caught: {e}')
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        print(f'ValidationError caught: {e}')
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        # Handle other unexpected exceptions
        print(f'Unexpected error caught: {e}')
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)            

@login_required(login_url="/login/")
def amazon_remittance(request):
    date_field = ['Create_Date', 'Closed_On']
    excluded_fields = ['id','payment_number', 'bol', 'backup_status', 'invoice_status', 'pod_status', 'validation_status', 'invalid_amount', 'valid_amount']
    #setup pagination
    data = Paginator(amazonremittance.objects.all(),10)
    page = request.GET.get('page')
    page_obj = data.get_page(page)
    fields = amazonremittance._meta.fields
    

    return render(request, 'promo_data.html', {'date_field':date_field, 'page_obj': page_obj, 'data': data, 'fields': fields})

@login_required(login_url="/login/")
def promo_backup(request):
    date_field = ['Receive_Date', 'Return_Date', 'Invoice_Day', 'Price_Protection_Day']
    #setup pagination
    data = Paginator(amazonpromo.objects.all(),10)
    page = request.GET.get('page')
    page_obj = data.get_page(page)
    fields = amazonpromo._meta.fields

    return render(request, 'amazon_promo_backup.html', {'page_obj': page_obj, 'data': data, 'fields': fields, 'date_field': date_field})

@login_required(login_url="/login/")
def amazonpo_data(request):
    
    #setup pagination
    data = Paginator(amazonpodata.objects.all(),20)
    page = request.GET.get('page')
    page_obj = data.get_page(page)
    fields = amazonpodata._meta.fields

    return render(request, 'po_data.html', {'page_obj': page_obj, 'data': data, 'fields': fields})

@login_required(login_url="/login/")
def amazon_contract(request):
    date_fields = ['AGREEMENT_END','INVOICE_DATE']
    rebate_fields =['REBATE_prct']
    #setup pagination
    data = Paginator(amazonagreement.objects.all(),20)
    page = request.GET.get('page')
    page_obj = data.get_page(page)
    fields = amazonagreement._meta.fields

    return render(request, 'amazon_contract_data.html', {'page_obj': page_obj, 'data': data, 'fields': fields, 'date_fields':date_fields})

@login_required(login_url="/login/")
def transform(request):

    if request.method == 'POST':
        form = TransformDataForm(request.POST)
        if form.is_valid():
            group_amazon_promo = form.cleaned_data['group_amazon_promo']
            po_calculation = form.cleaned_data['po_calculation']
            group_promo_validation = form.cleaned_data['group_promo_validation']
            promo_calculation = form.cleaned_data['promo_calculation']
            validation_rules = form.cleaned_data['validation_rules']
            


            if group_promo_validation:
                combine_promo_data(request)  # Run the data transformation function
                # test.delay()
                return redirect('transform_promotions_data')  # Redirect to the same page
            if group_amazon_promo:
                grouped_promos(request)  # Run the data transformation function
                return redirect('transform_promotions_data')  # Redirect to the same page
            
            if po_calculation:
                po_calculate(request)
                return redirect('transform_promotions_data')  # Redirect to the same page
            
            if promo_calculation:
                promo_calculate(request)
                return redirect('transform_promotions_data')  # Redirect to the same page
            
            if validation_rules:
                promo_rules(request)
                return redirect('transform_promotions_data')  # Redirect to the same page

    else:
        form = TransformDataForm()
    
    context = {
        'form': form,
    }
    return render(request, 'transform_promotions_data.html', context)

def promo_calculate(request):
        update_unique_promo_sql = """
            UPDATE promotions_amazonpromo
            SET Unique_promo = Purchase_Order || Promo_Asin
            WHERE Unique_promo IS NULL OR Unique_promo = '';
        """

        # if self.Unique_promo is None or self.Unique_promo =="":
        #     self.Unique_promo = self.Purchase_Order + self.Promo_Asin
        # else:
        #     self.Unique_promo = self.Unique_promo
        with connection.cursor() as cursor:
            cursor.execute(update_unique_promo_sql)
        return render(request, 'transform_promotions_data.html')

def grouped_promos(request):
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO promotions_PromoGroup(
                    Deduction_ref,
                    unique_promo,
                    Receive_Date,
                    Return_Date,
                    Invoice_Day,
                    Transaction_Type,
                    Quantity,
                    Net_Receipts,
                    Net_Receipts_Currency,
                    List_Price,
                    List_Price_Currency,
                    Rebate_In_Agreement_Currency,
                    Agreement_Currency,
                    Rebate_In_Purchase_Order_Currency,
                    Purchase_Order_Currency,
                    Purchase_Order,
                    Promo_Asin,
                    UPC,
                    EAN,
                    Manufacturer,
                    Distributor,
                    Product_Group,
                    Category,
                    Subcategory,
                    Title,
                    Product_Description,
                    Binding,
                    Cost_Currency,
                    Old_Cost,
                    New_Cost,
                    Price_Protection_Agreement,
                    Price_Protection_Day,
                    Cost_Variance,
                    Invoice
                )
                SELECT
                    ap."Deduction_ref",
                    ap."Unique_promo",
                    ap."Receive_Date",
                    ap."Return_Date",
                    ap."Invoice_Day",
                    ap."Transaction_Type",
                    SUM(ap."Quantity") as "Quantity",
                    SUM(ap."Net_Receipts") as "Net_Receipts",
                    ap."Net_Receipts_Currency",
                    SUM(ap."List_Price") as "List_Price",
                    ap."List_Price_Currency",
                    SUM(ap."Rebate_In_Agreement_Currency") as "Rebate_In_Agreement_Currency",
                    ap."Agreement_Currency",
                    SUM(ap."Rebate_In_Purchase_Order_Currency") as "Rebate_In_Purchase_Order_Currency",
                    ap."Purchase_Order_Currency",
                    ap."Purchase_Order",
                    ap."Promo_Asin",
                    ap."UPC",
                    ap."EAN",
                    ap."Manufacturer",
                    ap."Distributor",
                    ap."Product_Group",
                    ap."Category",
                    ap."Subcategory",
                    ap."Title",
                    ap."Product_Description",
                    ap."Binding",
                    ap."Cost_Currency",
                    SUM(ap."Old_Cost") as "Old_Cost",
                    SUM(ap."New_Cost") as "New_Cost",
                    ap."Price_Protection_Agreement",
                    ap."Price_Protection_Day",
                    SUM(ap."Cost_Variance") as "Cost_Variance",
                    ap."Invoice"
                FROM promotions_amazonpromo ap            
                WHERE NOT EXISTS (
                    SELECT 1
                    FROM promotions_PromoGroup pg
                    WHERE pg.Deduction_ref = ap.Deduction_ref AND pg.unique_promo = ap.unique_promo
                )
                GROUP BY ap.Deduction_ref, ap.unique_promo
            """)
        
        # try:
            # Commit the transaction
            print("SQL statement executed successfully")
            transaction.commit()
            print("Transaction committed successfully")
    except IntegrityError:
            # Handle IntegrityError, which occurs if a duplicate entry is detected
            print("Duplicate entry detected.")
            error_message = 'Integrity error: Duplicate entry detected.'
            # print("Duplicate entry detected.")
            return render(request, 'transform_promotions_data.html', {'error_message': error_message})
            # pass
    except Exception as e:
        # Handle other unexpected exceptions
        error_message = f'An unexpected error occurred: {e}'
        return HttpResponseServerError(error_message)

    print("No exception occurred. Rendering the template.") 
    return render(request, 'transform_promotions_data.html')

def po_calculate(request):
    try:
        # SQL query for updating Unique_key
        update_unique_key_sql = """
            UPDATE promotions_amazonpodata
            SET Unique_key = PO || PO_ASIN
            WHERE Unique_key IS NULL OR Unique_key = '';
        """

        # SQL query for updating Eaches_Quantity
        update_eaches_quantity_sql = """
            UPDATE promotions_amazonpodata
            SET Eaches_Quantity = CASE
                WHEN (Eaches_Quantity IS NULL OR Eaches_Quantity = '') AND
                    (Case_Size IS NOT NULL AND Quantity_received IS NOT NULL) AND
                    (Case_Size != 0 OR Quantity_received != 0)
                THEN Case_Size * Quantity_received
                WHEN Eaches_Quantity != 0 THEN Eaches_Quantity
                ELSE 0
            END;

        """

        # Execute the raw SQL queries sequentially
        with connection.cursor() as cursor:
            cursor.execute(update_unique_key_sql)
            cursor.execute(update_eaches_quantity_sql)

        # Save the instance to persist the changes in the database
        transaction.commit()
        return redirect('transform_promotions_data')
    except Exception as e:
       
        print(f'An unexpected error occurred: {e}')
        error_message = f'An unexpected error occurred: {e}'
        return HttpResponseServerError(error_message)

    print("No exception occurred. Redirecting to 'transform_promotions_data'.")
    return redirect('transform_promotions_data')
    
def combine_promo_data(request):
    try:
        with connection.cursor() as cursor:
            cursor.execute("""
                INSERT INTO promotions_promovalidation (
                    payment_date,
                    invoice_number,
                    deduction_amount,
                    deduction_reason,
                    unique_promo,
                    promo_quantity,
                    net_receipts,
                    list_price,
                    purchase_order,
                    promo_asin,
                    rebate_in_agreement_currency,
                    unique_key,
                    podata_po,
                    podata_asin,
                    case_size,
                    quantity_received,
                    case_cost,
                    total_cost,
                    eaches_quantity,
                    agreement_num,
                    rebate_percent,
                    agreement_end
                )
                SELECT
                    COALESCE(ar."payment_date", ''),
                    COALESCE(ar."invoice_number", ''),
                    COALESCE(ar."deduction_amount", 0),
                    COALESCE(ar."deduction_reason", ''),
                    COALESCE(pg."unique_promo", ''),
                    COALESCE(pg."Quantity", 0),
                    COALESCE(pg."Net_Receipts", 0),
                    COALESCE(pg."List_Price", 0),
                    COALESCE(pg."Purchase_Order", ''),
                    COALESCE(pg."Promo_Asin", ''),
                    COALESCE(pg."Rebate_In_Agreement_Currency", 0),
                    COALESCE(apd."Unique_key", ''),
                    COALESCE(apd."PO", ''),
                    COALESCE(apd."PO_ASIN", ''),
                    COALESCE(apd."Case_Size", 0),
                    COALESCE(apd."Quantity_received", 0),
                    COALESCE(apd."Case_Cost", 0),
                    COALESCE(apd."Total_Cost", 0),
                    COALESCE(apd."Eaches_Quantity", 0),
                    COALESCE(aa."AGREEMENT_num", ''),
                    COALESCE(aa."REBATE_prct", 0),
                    COALESCE(aa."AGREEMENT_END", '')
                FROM promotions_PromoGroup pg
                LEFT JOIN (
                    SELECT "payment_date", "invoice_number", "deduction_amount", "deduction_reason"
                    FROM promotions_amazonremittance
                    WHERE ("invoice_number", "payment_date") IN (
                        SELECT "invoice_number", MAX("payment_date") AS "Max_payment_date"
                        FROM promotions_amazonremittance
                        GROUP BY "invoice_number"
                    )           
                ) AS ar ON pg."Deduction_ref" = ar."invoice_number"
                LEFT JOIN promotions_amazonpodata apd ON pg."unique_promo" = apd."Unique_key"
                LEFT JOIN promotions_amazonagreement aa ON SUBSTR(pg."Deduction_ref", 1, 12) = SUBSTR(aa."INVOICE_NUMBER", 1, 12)
                WHERE NOT EXISTS (
                    SELECT 1
                    FROM promotions_promovalidation apv
                    WHERE apv."unique_promo" = pg."unique_promo"
                    AND apv."invoice_number" = COALESCE(ar."invoice_number", '')
                );
            """)
        
        # try:
            # Commit the transaction
            transaction.commit()
            print("Transaction committed successfully")

    
    except IntegrityError:
        # Handle IntegrityError, which occurs if a duplicate entry is detected
        pass
        print("IntegrityError: Duplicate entry detected.")
        transaction.rollback()
        return render(request, 'transform_promotions_data.html')

    except Exception as e:
    # Handle other exceptions
        print(f"Unexpected error: {e}")
        transaction.rollback()
        return render(request, 'transform_promotions_data.html')

def promo_rules(request):
    # SQL query to update valid_rebate_rate
        update_unit_price = """
            UPDATE promotions_promovalidation
            SET unit_price = 
                CASE
                    WHEN net_receipts IS NOT NULL AND net_receipts <> 0 AND promo_quantity <> 0
                    THEN ROUND(net_receipts / promo_quantity, 2)
                    ELSE 0
                END
            WHERE unit_price IS NULL OR unit_price = '';
        """
        update_quantity_variance = """
            UPDATE promotions_promovalidation
            SET quantity_variance = 
                CASE
                    WHEN promo_quantity IS NOT NULL AND quantity_received IS NOT NULL AND promo_quantity <> 0
                    THEN promo_quantity - quantity_received
                    ELSE 0
                END
            WHERE quantity_variance IS NULL OR quantity_variance = '';
        """

        update_actual_netreceipts = """
            UPDATE promotions_promovalidation
            SET actual_netreceipts = 
                CASE
                    WHEN eaches_quantity IS NOT NULL
                    THEN eaches_quantity * unit_price
                    ELSE actual_netreceipts
                END
            WHERE actual_netreceipts IS NULL OR actual_netreceipts = '' OR actual_netreceipts = 0;
        """

        update_netreceipts_variance = """
            UPDATE promotions_promovalidation
            SET netreceipts_variance = 
                CASE
                    WHEN net_receipts IS NOT NULL AND actual_netreceipts IS NOT NULL AND net_receipts <> 0
                    THEN net_receipts - actual_netreceipts
                    ELSE netreceipts_variance
                END
            WHERE netreceipts_variance IS NULL OR netreceipts_variance = '' OR netreceipts_variance = 0;
        """


        update_valid_rebate_rate = """
            UPDATE promotions_promovalidation
            SET valid_rebate_rate = CASE
                WHEN valid_rebate_rate IS NULL OR valid_rebate_rate = ''
                    THEN CASE
                        WHEN net_receipts IS NOT NULL AND rebate_in_agreement_currency IS NOT NULL AND net_receipts != 0
                            THEN
                                CASE
                                    WHEN ROUND(rebate_in_agreement_currency / net_receipts, 3) <= ROUND(rebate_percent, 3)
                                        THEN 'Valid'
                                    ELSE 'Invalid'
                                END
                            ELSE 'Invalid (Division by zero)'
                        END
                ELSE valid_rebate_rate
            END;
        """
        update_valid_po = """
            UPDATE promotions_promovalidation
            SET valid_po = 
                CASE
                    WHEN valid_po IS NULL OR valid_po = ''
                    THEN
                        CASE
                            WHEN purchase_order = podata_po
                            THEN 'Valid'
                            ELSE 'Invalid'
                        END
                    ELSE valid_po
                END;
        """
        update_valid_sku = """
            UPDATE promotions_promovalidation
            SET valid_sku = 
                CASE
                    WHEN valid_sku IS NULL OR valid_sku = ''
                    THEN
                        CASE
                            WHEN promo_asin = podata_asin
                            THEN 'Valid'
                            ELSE 'Invalid'
                        END
                    ELSE valid_sku
                END;
        """
        update_qty_variance = """
            UPDATE promotions_promovalidation
            SET qty_variance = 
                CASE
                    WHEN qty_variance IS NULL OR qty_variance = ''
                    THEN
                        CASE
                            WHEN quantity_variance IS NOT NULL AND quantity_variance > 0
                            THEN 'Invalid'
                            ELSE 'Valid'
                        END
                    ELSE qty_variance
                END;
        """
        update_validation_status = """
            UPDATE promotions_promovalidation
            SET validation_status = 
                CASE
                    WHEN validation_status IS NULL OR validation_status = ''
                    THEN
                        CASE
                            WHEN valid_rebate_rate = 'Invalid'
                            THEN 'Invalid'
                            WHEN qty_variance = 'Invalid'
                            THEN 'Invalid'
                            WHEN valid_po = 'Invalid'
                            THEN 'Invalid'
                            WHEN valid_sku = 'Invalid'
                            THEN 'Invalid'
                            ELSE 'Valid'
                        END
                    ELSE validation_status
                END;
        """
        update_detailed_reason = """
                UPDATE promotions_promovalidation
                SET detailed_reason =
                CASE
                    WHEN detailed_reason IS NULL OR detailed_reason = ''
                    THEN
                        CASE
                            WHEN valid_po = 'Invalid'
                            THEN 'PO deducted is not billed in the system'
                            WHEN valid_sku = 'Invalid'
                            THEN 'ASIN deducted is not billed on PO'
                            WHEN qty_variance = 'Invalid'
                            THEN 'Quantity deducted is higher than received quantity'
                            WHEN valid_rebate_rate = 'Invalid'
                            THEN 'Rebate percent taken is higher than rebate percent allowed in the contract'
                            ELSE ''
                        END
                    ELSE detailed_reason
                END;
        """
        update_invalid_amount = """
                UPDATE promotions_promovalidation
                SET invalid_amount = 
                CASE
                    WHEN invalid_amount IS NULL OR invalid_amount = ''
                    THEN
                        CASE
                            
                            WHEN valid_po = 'Invalid'
                            THEN rebate_in_agreement_currency
                            WHEN valid_sku = 'Invalid'
                            THEN rebate_in_agreement_currency
                            WHEN qty_variance = 'Invalid'
                            THEN netreceipts_variance * rebate_percent
                            WHEN valid_rebate_rate = 'Invalid'
                            THEN ((rebate_in_agreement_currency/net_receipts)-rebate_percent) * promo_quantity
                            ELSE 0
                        END
                    ELSE invalid_amount
                END;
        """

        with connection.cursor() as cursor:
            print("Executing update_unit_price")
            cursor.execute(update_unit_price)
            print("Update update_unit_price executed successfully")
            print("Executing update_quantity_variance")
            cursor.execute(update_quantity_variance)
            print("Update update_quantity_variance executed successfully")
            cursor.execute(update_actual_netreceipts)
            print("Update update_actual_netreceipt")
            cursor.execute(update_netreceipts_variance)
            print("Update update_netreceipts_variance")
            cursor.execute(update_valid_rebate_rate)
            print("Update update_valid_rebate_rate")
            cursor.execute(update_valid_po)
            print("Update update_valid_po")
            cursor.execute(update_valid_sku)
            print("Update update_valid_sku")
            cursor.execute(update_qty_variance)
            print("Update update_qty_variance")
            cursor.execute(update_validation_status)
            print("Update update_validation_status")
            cursor.execute(update_detailed_reason)
            print("Update update_detailed_reason")
            cursor.execute(update_invalid_amount)
            print("Update_invalid_amount")



        return redirect('transform_promotions_data')

def calculate_total_deductions(queryset):
    # Sum of REBATE IN AGREEMENT CURRENCY
    total_deductions = queryset.aggregate(Sum('rebate_in_agreement_currency'))['rebate_in_agreement_currency__sum'] or 0

    # Count the number of rows in the queryset
    num_rows = queryset.count()

    # Round off the result to two decimal places
    total_deductions = round(total_deductions, 2)
    print("Total Deductions:", total_deductions)
    print("Number of Rows:", num_rows)
    
    return total_deductions, num_rows

def calculate_total_invalid(queryset):
    # Sum of INVALID AMOUNT
    total_invalid = queryset.aggregate(Sum('invalid_amount'))['invalid_amount__sum'] or 0

    # Count the number of rows in the queryset
    # invalid_rows = queryset.count()
    invalid_rows = queryset.filter(validation_status="Invalid").count()

    # Round off the result to two decimal places
    total_invalid = round(total_invalid, 2)

    return total_invalid,invalid_rows

def calculate_total_valid(queryset):
    # Total Deductions - Total Invalid
    total_deductions, _ = calculate_total_deductions(queryset)    
    total_invalid, _ = calculate_total_invalid(queryset)
    # Count the number of valid rows in the queryset
    valid_rows = queryset.filter(validation_status="Valid").count()
    # Round off the result to two decimal places
    print("valid calculation")
    total_valid = round(total_deductions - total_invalid, 2)
    print(total_valid, "total valid")
    return total_valid, valid_rows


def calculate_rca_split(queryset):
    # Initialize sum variables for each reason
    po_not_billed_sum = Decimal(0.0)
    asin_not_on_po_sum = Decimal(0.0)
    quantity_issue_sum = Decimal(0.0)
    rebate_issue_sum = Decimal(0.0)

    # Iterate through the queryset and accumulate sums for each reason
    for item in queryset:
        detailed_reason = item.detailed_reason  
        invalid_amount = item.invalid_amount  
        
        # Ensure detailed_reason and invalid_amount are not None
        if detailed_reason is not None and invalid_amount is not None:
            if 'PO deducted is not billed in the system' in detailed_reason:
                po_not_billed_sum += invalid_amount
            elif 'ASIN deducted is not billed on PO' in detailed_reason:
                asin_not_on_po_sum += invalid_amount
            elif 'Quantity deducted is higher than received quantity' in detailed_reason:
                quantity_issue_sum += invalid_amount
            elif 'Rebate percent taken is higher than rebate percent allowed in the contract' in detailed_reason:
                rebate_issue_sum += invalid_amount

    # Convert Decimal sums to float before constructing the dictionary
    sums = {
        'po_not_billed': float(po_not_billed_sum),
        'asin_not_on_po': float(asin_not_on_po_sum),
        'quantity_issue': float(quantity_issue_sum),
        'rebate_issue': float(rebate_issue_sum),
    }

    return sums

@login_required(login_url="/login/")
def promo_validation(request):
    try:
        date_fields = ['payment_date', 'agreement_end']
        deduction_fields = ['payment_date','invoice_number','deduction_amount','deduction_reason']
        promo_fields = ['unique_promo','promo_quantity','net_receipts','list_price','purchase_order','promo_asin','rebate_in_agreement_currency']
        po_fields =['unique_key','podata_po','podata_asin','case_size','quantity_received','case_cost','total_cost','eaches_quantity']
        contract_fields = ['agreement_num','rebate_percent','agreement_end']
        calculated_fields = ['unit_price','quantity_variance','actual_netreceipts','netreceipts_variance']
        rules_fields = ['valid_rebate_rate','valid_po','valid_sku','price_variance','qty_variance','validation_status','invalid_amount','detailed_reason']
        currency_fields = ['net_receipts','original_amount','list_price','rebate_in_agreement_currency','case_cost','total_cost','unit_price',
        'actual_netreceipts','netreceipts_variance','invalid_amount']
        quantity_fields = ['promo_quantity','case_size','quantity_received','eaches_quantity','quantity_variance']
        rebate_fields =['rebate_percent']
        # Get the filter parameters from the query string
        search_query = request.GET.get('search', '')
        sort_by = request.GET.get('sort', 'unique_key')
        filter_field = request.GET.get('filter', '')  # Get the filter field value
        
        # Filter the PromoGroup objects based on the search query and filter field
        promo_groups = PromoValidation.objects.filter(
            Q(unique_key__icontains=search_query) |  # Apply search query filter
            Q(validation_status__icontains=search_query)    # Apply filter field filter
        )
        
        # Sort the PromoGroup objects based on the sort_by parameter
        promo_groups = promo_groups.order_by(sort_by)
        
        
        promo_groups = promo_groups.exclude(rebate_in_agreement_currency__lte=0)        
        # Calculate the required values for the summary
        # total_deductions = calculate_total_deductions(promo_groups)
        total_deductions, num_rows = calculate_total_deductions(promo_groups)
        total_invalid, invalid_rows= calculate_total_invalid(promo_groups)
        total_valid,valid_rows = calculate_total_valid(promo_groups)
        rca_split = calculate_rca_split(promo_groups) 
        print(total_deductions,total_invalid,total_valid,"SUMMARY")    

        
        # Pagination
        data = Paginator(promo_groups, 10)
        page = request.GET.get('page')
        page_obj = data.get_page(page)
        fields = PromoValidation._meta.fields

        # Define excluded fields
        excluded_fields = ['id', 'Invoice_Day', 'Return_Date', 'Transaction_Type', 'Transaction_Type', 'Net_Receipts_Currency',
                        'List_Price_Currency', 'Agreement_Currency', 'Purchase_Order_Currency', 'UPC', 'EAN',
                        'Manufacturer', 'Distributor', 'Product_Group', 'Category', 'Subcategory', 'Title',
                        'Product_Description', 'Binding', 'Cost_Currency', 'Old_Cost', 'New_Cost',
                        'Price_Protection_Agreement', 'Price_Protection_Day', 'Cost_Variance', 'Invoice', 'unique_promo',
                        'list_price', 'case_size', 'price_variance','unit_price','quantity_variance','actual_netreceipts','netreceipts_variance']

        if 'export' in request.GET:
            excel_data = []
            for promo_group in promo_groups:
                row = [getattr(promo_group, field.name) for field in fields if field.name not in excluded_fields]
                excel_data.append(row)

            headers = [field.verbose_name for field in fields if field.name not in excluded_fields]

            # Create a new workbook and get the active sheet
            workbook = Workbook()
            sheet = workbook.active

            # Write headers to the first row
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num).value = header

            # Write data to the remaining rows
            for row_num, row in enumerate(excel_data, 2):
                for col_num, value in enumerate(row, 1):
                    sheet.cell(row=row_num, column=col_num).value = value

            # Set the appropriate response headers
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=promo_validation_data.xlsx'

            # Save the workbook to the response
            workbook.save(response)

            return response


        return render(request, 'promo_validation_engine.html', {'page_obj': page_obj, 'data': data, 'fields': fields,
                                                        'search_query': search_query, 'sort_by': sort_by,
                                                        'filter_field': filter_field, 'excluded_fields': excluded_fields,
                                                        'date_fields': date_fields, 'deduction_fields':deduction_fields,
                                                        'promo_fields': promo_fields, 'po_fields': po_fields, 'contract_fields':contract_fields,
                                                        'calculated_fields':calculated_fields, 'rules_fields':rules_fields,
                                                        'currency_fields':currency_fields, 'quantity_fields':quantity_fields, 'rebate_fields':rebate_fields,
                                                        'total_deductions': total_deductions,'num_rows': num_rows,'total_invalid': total_invalid,'invalid_rows':invalid_rows,
                                                        'total_valid': total_valid,'valid_rows':valid_rows,'rca_split': rca_split,
                                                            })
    
    except Exception as e:
        # Handle other unexpected exceptions
        print(f'An unexpected error occurred: {e}')
        error_message = f'An unexpected error occurred: {e}'
        return HttpResponseServerError(error_message)

@login_required(login_url="/login/")
def delete_data(request):
    if request.method == 'POST':
        form = DeleteDataForm(request.POST)
        if form.is_valid():
            
            delete_remit = form.cleaned_data['delete_remit']
            delete_promoagreement = form.cleaned_data['delete_promoagreement']
            delete_promobackup = form.cleaned_data['delete_promobackup']
            delete_promo_group = form.cleaned_data['delete_promo_group']
            delete_podata = form.cleaned_data['delete_podata']
            delete_promo_validation = form.cleaned_data['delete_promo_validation']
            
            if delete_remit:
                amazonremittance.objects.all().delete()
                
            if delete_promoagreement:
                amazonagreement.objects.all().delete()
                
            if delete_promobackup:
                amazonpromo.objects.all().delete()
            
            if delete_promo_group:
                PromoGroup.objects.all().delete()
            
            if delete_podata:
                amazonpodata.objects.all().delete()

            if delete_promo_validation:
                PromoValidation.objects.all().delete()

    else:
        form = DeleteDataForm()
    
    context = {
        'form': form,
    }

    return render(request, 'delete_data.html', context)




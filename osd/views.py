# pylint: disable=no-member
from django.template.loader import render_to_string
import io
from main.utils import *
from main import models as mm
from django.db import connection
from django.db import IntegrityError
from openpyxl import Workbook
import time
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import condition
from decimal import Decimal
from django.core.exceptions import ValidationError
from django.db.models import Q
from django.db import transaction
from django.db.models.functions import Concat
from django.db.models import Sum, F, Value, Avg, Max
from django.db.models import Count, OuterRef, Subquery
from datetime import datetime
from .models import *
from .forms import *
from django.shortcuts import render, HttpResponse, redirect
from django.contrib import messages
import pandas as pd
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.template import loader
from dateutil import parser
from pathlib import Path
from django.core.paginator import Paginator
from django.shortcuts import render
import logging
logger = logging.getLogger(__name__)
from main import views as mv


@login_required(login_url="/login/")
def deductions_upload(request, excel_file, context):
    try:
        if excel_file:
            # Load the Excel data into a pandas DataFrame
            df = pd.read_csv(excel_file, dtype={'id': str, 'invoice_number': str})
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df.shape[0]
            instances = []

            # Iterate over each row in the DataFrame and validate the data
            for _, row in df.iterrows():
                try:

                    ids = row['id']
                    customer_name = row['customer_name'] if 'customer_name' in row else ''
                    customer_account = row['customer_account'] if 'customer_account' in row else ''
                    standard_customer = row['standard_customer'] if 'standard_customer' in row else ''
                    deduction_reference = row['deduction_reference'] if 'deduction_reference' in row else ''
                    invoice_number = row['invoice_number']
                    deduction_date = row['deduction_date'] if 'deduction_date' in row else None
                    if pd.isnull(deduction_date):
                        deduction_date = None
                    else:
                        deduction_date = parse_date(row['deduction_date'])
                    deduction_amount = row['deduction_amount']
                    if deduction_amount is None or deduction_amount == '':
                        deduction_amount = 0
                    else:
                        deduction_amount = convert_to_decimal_V2(
                            row['deduction_amount'])
                    deduction_reason = row['deduction_reason'] if 'deduction_reason' in row else ''
                    payment_number = row['payment_number']
                    backup_status = row['backup_status'] if 'backup_status' in row else ''
                    invoice_status = row['invoice_status'] if 'invoice_status' in row else ''
                    bol = row['bol'] if 'bol' in row else ''
                    pod_status = row['pod_status'] if 'pod_status' in row else ''
                    invalid_amount = clean_value(row['invalid_amount'], 0.0)
                    valid_amount = clean_value(row['valid_amount'], 0.0)
                    priority = row['priority'] if 'priority' in row else ''
                    # Check if the record already exists in the model
                    if deduction_data.objects.filter(invoice_number=invoice_number, payment_number=payment_number, deduction_amount=deduction_amount).exists():
                        continue

                    # Create an instance of the model
                    instance = deduction_data(
                        ids=ids,
                        customer_name=customer_name,
                        customer_account=customer_account,
                        standard_customer=standard_customer,
                        deduction_reference=deduction_reference,
                        invoice_number=invoice_number,
                        deduction_date=deduction_date,
                        deduction_amount=deduction_amount,
                        deduction_reason=deduction_reason,
                        payment_number=payment_number,
                        backup_status=backup_status,
                        invoice_status=invoice_status,
                        bol=bol,
                        pod_status=pod_status,
                        invalid_amount=invalid_amount,
                        valid_amount=valid_amount,
                        priority = priority
                    )
                    instances.append(instance)

                except ValueError as e:
                    print(f"ValueError for row {row}: {e}")
                except Exception as e:
                    print(f"Unexpected error for row {row}: {e}")

            # Bulk insert the instances
            deduction_data.objects.bulk_create(instances)
            end_time = timezone.now()
            total_time = end_time - start_time

            # Update upload status
            current_status = 'completed'
            request.session['upload_status'] = 'success'

            # Log the upload details
            # uploaded_csv = mm.UploadedCSV.objects.create(
            #     user=request.user,
            #     file=excel_file,
            #     num_rows=df.shape[0],
            #     start_time=start_time,
            #     end_time=end_time,
            #     total_time=total_time,
            #     current_status=current_status
            # )
            return redirect('backup_upload')

        else:
            context['error'] = 'No Excel file was uploaded.'

    except pd.errors.EmptyDataError:
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)


@login_required(login_url="/login/")
def deductions_view(request):
    date_field = []
    excluded_fields = ['id','payment_number', 'bol', 'backup_status', 'invoice_status', 'pod_status', 'validation_status', 'invalid_amount', 'valid_amount']
    
    # Fetch records where validation_status is not None
    data = deduction_data.objects.all()  # Filter out None values for validation_status
    
    # Setup pagination
    paginator = Paginator(data, 10)
    page = request.GET.get('page')
    page_obj = paginator.get_page(page)

    # Get field names excluding certain fields
    fields = deduction_data._meta.fields
    field_names = [field.name for field in fields if field.name not in excluded_fields]
    
    return render(request, 'osd_deductions_view.html', {
        'date_field': date_field,
        'page_obj': page_obj,
        'data': data,
        'field_names': field_names
    })

@login_required(login_url="/login/")
def backup_upload(request, excel_file, context):
    try:
        if excel_file:
            # Load the Excel data into a pandas DataFrame
            df = pd.read_csv(excel_file, dtype={'id': str, 'invoice_number': str,'sku': str})
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df.shape[0]
            instances = []

            # Iterate over each row in the DataFrame and validate the data
            for _, row in df.iterrows():
                try:

                    ids = row['id']
                    invoice_number = row['invoice_number'] if 'invoice_number' in row else ''
                    sku = row['sku'] if 'sku' in row else ''
                    deducted_amount = row['deducted_amount']
                    if deducted_amount is None or deducted_amount == '':
                        deducted_amount = 0.0
                    else:
                        deducted_amount = convert_to_decimal_V2(
                            deducted_amount)
                        if deducted_amount is None or deducted_amount != deducted_amount:
                            deducted_amount = 0.0
                    deducted_qty = row['deducted_qty']
                    if deducted_qty is None or deducted_qty == '':
                        deducted_qty = 0
                    else:
                        deducted_qty = convert_to_decimal_V2(deducted_qty)
                        if deducted_qty is None or deducted_qty != deducted_qty:
                            deducted_qty = 0
                    deducted_price_per_qty = row['deducted_price_per_qty']
                    if deducted_price_per_qty is None or deducted_price_per_qty == '':
                        deducted_price_per_qty = 0.0
                    else:
                        deducted_price_per_qty = convert_to_decimal_V2(
                            deducted_price_per_qty)
                        if deducted_price_per_qty is None or deducted_price_per_qty != deducted_price_per_qty:
                            deducted_price_per_qty = 0.0

                    deduction_date = row['deduction_date'] if 'deduction_date' in row else None
                    if pd.isnull(deduction_date):
                        deduction_date = None
                    else:
                        deduction_date = parse_date(row['deduction_date'])

                    invoice_location = row['invoice_location'] if 'invoice_location' in row else ''
                    reason_code = row['reason_code'] if 'reason_code' in row else ''
                    sub_reason_code = row['sub_reason_code'] if 'sub_reason_code' in row else ''
                    deduction_reason = row['deduction_reason'] if 'deduction_reason' in row else ''

                    # Check if the record already exists in the model
                    if backup_data.objects.filter(invoice_number=invoice_number, sku=sku, deducted_amount=deducted_amount).exists():
                        continue

                    # Create an instance of the model
                    instance = backup_data(
                        ids=ids,
                        invoice_number=invoice_number,
                        sku=sku,
                        deducted_amount=deducted_amount,
                        deducted_qty=deducted_qty,
                        deducted_price_per_qty=deducted_price_per_qty,
                        deduction_date=deduction_date,
                        invoice_location=invoice_location,
                        reason_code=reason_code,
                        sub_reason_code=sub_reason_code,
                        deduction_reason=deduction_reason
                    )
                    instances.append(instance)

                except ValueError as e:
                    print(f"ValueError for row {row}: {e}")
                except Exception as e:
                    print(f"Unexpected error for row {row}: {e}")

            # Bulk insert the instances
            # backup_data.objects.bulk_create(instances)
            if instances:
                backup_data.objects.bulk_create(instances)
            end_time = timezone.now()
            total_time = end_time - start_time

            # Update upload status
            current_status = 'completed'
            request.session['upload_status'] = 'success'

            # Log the upload details
            # uploaded_csv = mm.UploadedCSV.objects.create(
            #     user=request.user,
            #     file=excel_file,
            #     num_rows=df.shape[0],
            #     start_time=start_time,
            #     end_time=end_time,
            #     total_time=total_time,
            #     current_status=current_status
            # )
            return redirect('backup_upload')

        else:
            context['error'] = 'No Excel file was uploaded.'

    except pd.errors.EmptyDataError:
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)


@login_required(login_url="/login/")
def backup_view(request):
    date_field = ['deduction_date']
    excluded_fields = ['id','ids', 'invoice_location', 'sub_reason_code','reason_code']
    # Setup pagination
    data = Paginator(backup_data.objects.all(), 10)
    page = request.GET.get('page')
    page_obj = data.get_page(page)
    fields = backup_data._meta.fields
    field_names = [
        field.name for field in fields if field.name not in excluded_fields]
    print(field_names)

    return render(request, 'conagra_backup.html', {'date_field': date_field, 'page_obj': page_obj, 'data': data, 'field_names': field_names})


@login_required(login_url="/login/")
def invoice_upload(request, excel_file=None, context={}):

    try:
        if excel_file:
            # Load the Excel data into a pandas DataFrame
            df = pd.read_csv(excel_file, dtype={'invoice_number': str, 'sku': str})
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df.shape[0]
            instances = []

            # Define a function to convert NaN to default values
            def get_value(value, default=0):
                return default if pd.isna(value) else value

            # Iterate over each row in the DataFrame and validate the data
            for _, row in df.iterrows():
                try:
                    # Extract fields as per the model
                    inv_sku = row['concat'] if 'concat' in row else ''
                    invoice_number = row['invoice_number'] if 'invoice_number' in row else ''
                    sku = row['sku'] if 'sku' in row else ''
                    billed_qty = clean_value(row['billed_qty'], 0)
                    gross_price = row['gross_price']
                    if gross_price is None or gross_price == '':
                        gross_price = 0.0
                    else:
                        gross_price = convert_to_decimal_V2(gross_price)
                        if gross_price is None or gross_price != gross_price:
                            gross_price = 0.0
                    
                    oi_deal = clean_value(row['oi_deal'], 0.0)
                    promo_allowance = clean_value(row['promo_allowance'], 0.0)
                    cash_discount = clean_value(row['cash_discount'], 0.0)
                    freight_amount = clean_value(row['freight_amount'], 0.0)
                    others = clean_value(row['others'], 0.0)
                    net_price = row['net_price']
                    if net_price is None or net_price == '':
                        net_price = 0.0
                    else:
                        net_price = convert_to_decimal_V2(net_price)
                        if net_price is None or net_price != net_price:
                            net_price = 0.0
                    gross_price_per_qty = clean_value(
                        row['gross_price_per_qty'], 0.0)
                    net_price_per_qty = clean_value(
                        row['net_price_per_qty'], 0.0)
                    order_number = row['order_number'] if 'order_number' in row else ''
                    bol = row['bol'] if 'bol' in row else ''
                    carrier = row['carrier'] if 'carrier' in row else ''
                    freight_type = row['freight_type'] if 'freight_type' in row else ''

                    # Check if the record already exists in the model
                    if invoice_data.objects.filter(inv_sku=inv_sku, order_number=order_number).exists():
                        continue

                    # Create an instance of the model
                    instance = invoice_data(
                        inv_sku=inv_sku,
                        invoice_number=invoice_number,
                        sku=sku,
                        billed_qty=billed_qty,
                        gross_price=gross_price,
                        oi_deal=oi_deal,
                        promo_allowance=promo_allowance,
                        cash_discount=cash_discount,
                        freight_amount = freight_amount,
                        others=others,
                        net_price=net_price,
                        gross_price_per_qty=gross_price_per_qty,
                        net_price_per_qty=net_price_per_qty,
                        order_number=order_number,
                        bol=bol,
                        carrier=carrier,
                        freight_type=freight_type
                    )
                    instances.append(instance)

                except ValueError as e:
                    print(f"ValueError for row {row}: {e}")
                except Exception as e:
                    print(f"Unexpected error for row {row}: {e}")

            # Bulk insert the instances
            if instances:
                invoice_data.objects.bulk_create(instances)

            end_time = timezone.now()
            total_time = end_time - start_time

            # Update upload status
            current_status = 'completed'
            request.session['upload_status'] = 'success'

            # Log the upload details
            # uploaded_csv = mm.UploadedCSV.objects.create(
            #     user=request.user,
            #     file=excel_file,
            #     num_rows=df.shape[0],
            #     start_time=start_time,
            #     end_time=end_time,
            #     total_time=total_time,
            #     current_status=current_status
            # )
            return redirect('backup_upload')

        else:
            context['error'] = 'No Excel file was uploaded.'

    except pd.errors.EmptyDataError:
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)


def pod_detail_upload(request, excel_file=None, context={}):
    try:
        if excel_file:
            # Load the Excel data into a pandas DataFrame
            df = pd.read_csv(excel_file)
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df.shape[0]
            instances = []

            # Define a function to convert NaN to default values
            def clean_value(value, default=None):
                if pd.isna(value):
                    return default
                return value

            # Iterate over each row in the DataFrame and validate the data
            for _, row in df.iterrows():
                try:
                    # Extract fields as per the model
                    inv_sku = row['concat'] if 'concat' in row else ''
                    # file paths
                    pod_file = row['pod_file'] if 'pod_file' in row else 'NoFile.pdf'
                    order_number = row['order_number'] if 'order_number' in row else ''
                    invoice_number = row['invoice_number'] if 'invoice_number' in row else ''
                    bol = row['bol'] if 'bol' in row else ''
                    sku = row['sku'] if 'sku' in row else ''
                    shortage = clean_value(row['shortage'], 0)
                    damage = clean_value(row['damage'], 0)
                    returns = clean_value(row['return'], 0)
                    overage = clean_value(row['overage'], 0)
                    net_shortage = clean_value(row['net_shortage'], 0)
                    customer_sign = row['customer_sign'] if 'customer_sign' in row else ''
                    carrier_sign = row['carrier_sign'] if 'carrier_sign' in row else ''
                    subject_to_count = row['subject_to_count'] if 'subject_to_count' in row else False
                    pod_found = row['pod_found'] if 'pod_found' in row else False

                    # Check if the record already exists in the model
                    if pod_detail.objects.filter(inv_sku=inv_sku, order_number=order_number).exists():
                        continue

                    # Create an instance of the model
                    instance = pod_detail(
                        inv_sku=inv_sku,
                        pod_file=pod_file,
                        order_number=order_number,
                        invoice_number=invoice_number,
                        bol=bol,
                        sku=sku,
                        shortage=shortage,
                        damage=damage,
                        returns=returns,
                        overage=overage,
                        net_shortage=net_shortage,
                        customer_sign=customer_sign,
                        carrier_sign=carrier_sign,
                        subject_to_count=subject_to_count,
                        pod_found=pod_found
                    )
                    instances.append(instance)

                except ValueError as e:
                    print(f"ValueError for row {row}: {e}")
                except Exception as e:
                    print(f"Unexpected error for row {row}: {e}")

            # Bulk insert the instances
            if instances:
                pod_detail.objects.bulk_create(instances)

            end_time = timezone.now()
            total_time = end_time - start_time

            # Update upload status
            current_status = 'completed'
            request.session['upload_status'] = 'success'

            # Log the upload details
            # Assuming mm.UploadedCSV is a model for logging
            uploaded_csv = mm.UploadedCSV.objects.create(
                user=request.user,
                file=excel_file,
                num_rows=df.shape[0],
                start_time=start_time,
                end_time=end_time,
                total_time=total_time,
                current_status=current_status
            )
            return redirect('backup_upload')

        else:
            context['error'] = 'No Excel file was uploaded.'

    except pd.errors.EmptyDataError:
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)


def pdf_list(request):

    data = Paginator(pod_detail.objects.all(), 10)
    page = request.GET.get('page')
    page_obj = data.get_page(page)
    fields = pod_detail._meta.fields

    return render(request, 'pdf_list.html', {'page_obj': page_obj, 'data': data, 'fields': fields})


def update_pod_details(request, pod_id):
    pod_instance = get_object_or_404(pod_detail, pk=pod_id)
    invoice_number = pod_instance.invoice_number  # Assuming there's an invoice_number field in the model

    if request.method == 'POST':
        form = PODDetailsForm(request.POST, request.FILES, instance=pod_instance)
        if form.is_valid():
            # Check if a new file is provided in the form
            new_pdf_file = form.cleaned_data.get('pod_file')
            if new_pdf_file:
                # Update pod_file and title for all rows with the same invoice number
                pod_records = pod_detail.objects.filter(invoice_number=invoice_number)
                for pod in pod_records:
                    pod.pod_file = new_pdf_file
                    pod.Title = new_pdf_file.name
                    pod.save()

            success_message = 'POD updated successfully for all rows with the same invoice number.'
            return render(request, 'pod_update.html', {'form': form, 'success_message': success_message})
    else:
        form = PODDetailsForm(instance=pod_instance)

    return render(request, 'pod_update.html', {'form': form})



def upload_pod(request):
    if request.method == 'POST':
        form = PODUploadForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                uploaded_file_name = request.FILES['pod_file'].name
                form_instance = form.save(commit=False)
                form_instance.Title = uploaded_file_name
                form_instance.save()
                return redirect('backup_upload')
            except Exception as e:
                logger.error(f"Error uploading file: {e}")
        else:
            logger.error(f"Form validation errors: {form.errors}")
    else:
        form = PODUploadForm()
    return render(request, 'upload_pod.html', {'form': form})


def transform_osd_data(request):
    if request.method == 'POST':
        form = TransformDataForm(request.POST)
        if form.is_valid():
            osd_validation = form.cleaned_data['osd_validation']
            osd_calculations = form.cleaned_data['osd_calculations']
            deductions_calculations = form.cleaned_data['deductions_calculations']
            if osd_validation:
                combine_osd_data(request)
                return redirect('transform_osd_data')
            if osd_calculations:
                update_calculated_fields(request)
                return redirect('transform_osd_data')
            if deductions_calculations:
                update_deductions_data(request)
                return redirect('transform_osd_data')
    else:
        form = TransformDataForm()
    context = {
        'form': form,
    }
    print("Rendering transform_osd_data.html...")
    return render(request, 'transform_osd_data.html', context)



def combine_osd_data(request):
    validation.objects.all().delete()
    try:
        print("Attempting to combine OSD data...")
        with connection.cursor() as cursor:
            sql_query = """
                INSERT INTO osd_validation (ids, standard_customer, deduction_reference, invoice_number, sku, deducted_amount,
                                        deducted_qty, deducted_price_per_qty, deduction_date, deduction_reason, billed_qty,
                                        gross_price_per_qty, net_price_per_qty, carrier, shortage, damage, returns, overage, 
                                        net_shortage, customer_sign, carrier_sign, subject_to_count)
                SELECT DISTINCT
                    COALESCE(dd."ids", '') ,
                    COALESCE(dd."standard_customer", '') ,
                    COALESCE(dd."deduction_reference", '') ,
                    COALESCE(dd."invoice_number", ''), 
                    COALESCE(bd."sku", '') ,
                    COALESCE(bd."deducted_amount", 0) ,
                    COALESCE(bd."deducted_qty", 0) ,
                    COALESCE(bd."deducted_price_per_qty", 0) ,
                    COALESCE(bd."deduction_date", '') ,
                    COALESCE(bd."deduction_reason", '') ,
                    COALESCE(id."billed_qty", 0) AS "billed_qty" ,
                    COALESCE(id."gross_price_per_qty", 0) AS "gross_price_per_qty",
                    COALESCE(id."net_price_per_qty", 0) AS "net_price_per_qty",
                    COALESCE(id."carrier", '') ,
                    COALESCE(pd."shortage", 0) ,
                    COALESCE(pd."damage", 0) ,
                    COALESCE(pd."returns", 0) ,
                    COALESCE(pd."overage", 0) ,
                    COALESCE(pd."net_shortage", 0) ,
                    COALESCE(pd."customer_sign", 'Yes') ,
                    COALESCE(pd."carrier_sign", 'Yes') ,
                    COALESCE(pd."subject_to_count", 'Yes') 
                FROM
                    osd_backup_data bd
                    LEFT JOIN osd_deduction_data dd ON bd.ids = dd.ids
                    LEFT JOIN main_invoice_data id ON bd.invoice_number = id.invoice_number AND bd.sku = id.sku
                    LEFT JOIN osd_pod_detail pd ON bd.sku = pd.sku AND bd.invoice_number = pd.invoice_number 
            """
            print("Executing SQL query...")
            cursor.execute(sql_query)
    except IntegrityError as e:
        error_message = f"Integrity Error: {e}"
        print(error_message)
        return render(request, 'transform_osd_data.html', {'error_message': error_message})
    except Exception as e:
        print(f"Error: {e}")
        error_message = f"Database Error: {e}"
        return render(request, 'transform_osd_data.html', {'error_message': error_message})

    print("combine_osd_data executed successfully.")
    return render(request, 'transform_osd_data.html', {'success_message': 'OSD data combined successfully.'})


def update_calculated_fields(self):
    # SQL queries to update calculated fields
    update_missing_data = """
            UPDATE osd_validation
            SET missing_data = CASE
                WHEN sku = 0 OR sku = '' THEN 'Missing Backup'
                WHEN billed_qty = 0 OR billed_qty = '' THEN 'Missing Invoice data'
                ELSE 'Data Complete'
            END
            WHERE missing_data IS NULL;
        """

    update_no_shortage_in_pod = """
            UPDATE osd_validation
            SET no_shortage_in_pod = CASE
                WHEN net_shortage = 0 AND billed_qty <> 0 THEN 'Invalid'
                ELSE 'Valid'
            END
            WHERE no_shortage_in_pod IS NULL;
        """

    update_invalid_amt_1 = """
            UPDATE osd_validation
            SET invalid_amt_1 = CASE
                WHEN no_shortage_in_pod = 'Invalid' THEN deducted_amount
                ELSE 0
            END
            WHERE invalid_amt_1 IS NULL;
        """

    update_partial_shortage_in_pod = """
            UPDATE osd_validation
                SET partial_shortage_in_pod = CASE
                    WHEN net_shortage > 0 
                        AND net_shortage < COALESCE(deducted_qty, 0)
                        AND COALESCE(billed_qty, 0) <> 0 THEN 'Invalid'
                    ELSE 'Valid'
                END
                WHERE partial_shortage_in_pod IS NULL;
        """

    update_invalid_amt_2 = """
            UPDATE osd_validation
            SET invalid_amt_2 = CASE
                WHEN partial_shortage_in_pod = 'Invalid' THEN (deducted_qty - net_shortage) * deducted_price_per_qty
                ELSE 0
            END
            WHERE invalid_amt_2 IS NULL;
        """

    update_deducted_sku_is_not_invoiced = """
            UPDATE osd_validation
            SET deducted_sku_is_not_invoiced = CASE
                WHEN billed_qty = '' OR billed_qty = 0 THEN 'Invalid'
                ELSE 'Valid'
            END
            WHERE deducted_sku_is_not_invoiced IS NULL;
        """

    update_invalid_amt_3 = """
            UPDATE osd_validation
            SET invalid_amt_3 = CASE
                WHEN deducted_sku_is_not_invoiced = 'Invalid' THEN deducted_amount
                ELSE 0
            END
            WHERE invalid_amt_3 IS NULL;
        """

    update_pricing_variance = """
            UPDATE osd_validation
            SET pricing_variance = CASE
                WHEN billed_qty = '' OR billed_qty = 0 THEN 'SKU is not Invoiced'
                WHEN deducted_price_per_qty > net_price_per_qty AND no_shortage_in_pod = 'Valid' AND partial_shortage_in_pod = 'Valid' THEN 'Invalid'
                ELSE 'Valid'
            END
            WHERE pricing_variance IS NULL;
        """

    update_invalid_amt_4 = """
            UPDATE osd_validation
            SET invalid_amt_4 = CASE
                WHEN pricing_variance = 'Invalid' THEN (deducted_price_per_qty - net_price_per_qty) * deducted_qty
                ELSE 0
            END
            WHERE invalid_amt_4 IS NULL;
        """
    update_validation_status = """
            UPDATE osd_validation
            SET validation_status = 
                CASE
                    WHEN customer_sign = 'No' THEN 'Pending Validation'
                    WHEN carrier_sign = 'No' THEN 'Pending Validation'
                    WHEN subject_to_count = 'Yes' THEN 'Pending Validation'
                    WHEN no_shortage_in_pod = 'Invalid' THEN 'Invalid'
                    WHEN partial_shortage_in_pod = 'Invalid' THEN 'Partially Invalid'
                    WHEN deducted_sku_is_not_invoiced = 'Invalid' THEN 'Invalid'
                    WHEN pricing_variance = 'Invalid' THEN 'Invalid'
                    ELSE 'Valid'
                END
            WHERE validation_status IS NULL OR validation_status = '';

        """
    update_invalid_reason = """
                UPDATE osd_validation
                SET invalid_reason =
                    CASE
                        WHEN invalid_reason IS NULL OR invalid_reason = ''
                        THEN
                            CASE
                                WHEN validation_status = 'Pending Validation' AND customer_sign = 'No'
                                THEN 'Customer sign is missing on POD'
                                WHEN validation_status = 'Pending Validation' AND carrier_sign = 'No'
                                THEN 'Carrier sign is missing on POD'
                                WHEN validation_status = 'Pending Validation' AND subject_to_count = 'Yes'
                                THEN 'POD is subject to count'
                                WHEN missing_data = 'Missing Backup'
                                THEN 'Backup is not uploaded'
                                WHEN no_shortage_in_pod = 'Invalid'
                                THEN 'POD does not support shortages'
                                WHEN partial_shortage_in_pod = 'Invalid'
                                THEN 'POD supports partial shortages'
                                WHEN deducted_sku_is_not_invoiced = 'Invalid'
                                THEN 'Deducted SKU is not invoiced'
                                WHEN pricing_variance = 'Invalid'
                                THEN 'Deduction is taken at higher price than billed price'
                                ELSE 'Valid'
                            END
                        ELSE invalid_reason
                    END;
        """
    update_invalid_amt = """
                UPDATE osd_validation
                SET invalid_amount = CASE
                    WHEN validation_status = 'Invalid' AND invalid_amt_1 IS NOT NULL THEN COALESCE(invalid_amt_1, 0) + COALESCE(invalid_amt_2, 0) + COALESCE(invalid_amt_3, 0) + COALESCE(invalid_amt_4, 0)
                    WHEN validation_status = 'Partially Invalid' AND invalid_amt_1 IS NOT NULL THEN COALESCE(invalid_amt_1, 0) + COALESCE(invalid_amt_2, 0) + COALESCE(invalid_amt_3, 0) + COALESCE(invalid_amt_4, 0)
                    ELSE 0
                END
                WHERE invalid_amount IS NULL OR invalid_amount = '';

        """

    update_valid_amt = """
                UPDATE osd_validation
                SET valid_amount = CASE
                    WHEN validation_status <> 'Pending Validation' THEN COALESCE(deducted_amount, 0) - COALESCE(invalid_amount, 0)
                    ELSE 0
                END
                WHERE valid_amount IS NULL OR valid_amount = '';
        """

    update_rca = """
                UPDATE osd_validation
                SET product_substitution = 'Data not available',
                    combined_shipment = 'Data not available',
                    order_split = 'Data not available',
                    load_sequencing = 'Data not available',
                    unit_of_measurement = 'Data not available',
                    deducted_at_higher_price = 'Data not available'
                

        """
    with connection.cursor() as cursor:
        cursor.execute(update_missing_data)
        cursor.execute(update_no_shortage_in_pod)
        cursor.execute(update_invalid_amt_1)
        cursor.execute(update_partial_shortage_in_pod)
        cursor.execute(update_invalid_amt_2)
        cursor.execute(update_deducted_sku_is_not_invoiced)
        cursor.execute(update_invalid_amt_3)
        cursor.execute(update_pricing_variance)
        cursor.execute(update_invalid_amt_4)
        cursor.execute(update_validation_status)
        cursor.execute(update_invalid_reason)
        cursor.execute(update_invalid_amt)
        cursor.execute(update_valid_amt)
        cursor.execute(update_rca)

    return redirect('transform_osd_data')


@login_required(login_url="/login/")
def osd_validation_view(request):
    # Define the fields
    date_fields = ['deduction_date']
    deduction_fields = ['standard_customer', 'deduction_reference', 'invoice_number' ]
    backup_fields = ['sku','deducted_amount', 'deducted_qty', 'deducted_price_per_qty', 'deduction_date',
                    'deduction_reason']
    invoice_fields = ['billed_qty', 'gross_price_per_qty', 'net_price_per_qty', 'carrier']
    POD_fields = ['shortage', 'damage', 'returns', 'overage', 'net_shortage', ]
    calculated_fields = ['no_shortage_in_pod', 'partial_shortage_in_pod', 
                         'deducted_sku_is_not_invoiced', 'pricing_variance', 
                         'validation_status', 'invalid_amount', 'valid_amount', 'invalid_reason',]
    rca_fields = ['product_substitution', 'combined_shipment', 'order_split', 'load_sequencing', 'unit_of_measurement', 'deducted_at_higher_price']

    # Get the filter parameters from the query string
    search_query = request.GET.get('search', '')
    # Initialize sort_by variable
    sort_by = request.GET.get('sort', 'invoice_number')
    filter_field = request.GET.get('filter', '')  # Get the filter field value

    # Filter the validation objects based on the search query
    osd_groups = validation.objects.filter(
        Q(invoice_number__icontains=search_query)  # Apply search query filter
    )

    # Sort the validation objects based on the sort_by parameter
    try:
        osd_groups = osd_groups.order_by(sort_by)
    except TypeError as e:
        print("Error during sorting:", e)

    # Calculate the required values for the summary
    total_deductions, num_rows = calculate_total_deductions(osd_groups)
    total_invalid, invalid_rows = calculate_total_invalid(osd_groups)
    total_valid, valid_rows = calculate_total_valid(osd_groups)
    rca_split = calculate_rca_split(osd_groups)

    # Pagination
    data = Paginator(osd_groups, 10)
    page = request.GET.get('page')
    page_obj = data.get_page(page)
    fields = validation._meta.fields

    # Define excluded fields
    excluded_fields = ['id','ids','dedution_reference', 'invalid_amt_1',  'invalid_amt_2',
                          'invalid_amt_3',  'invalid_amt_4',  'customer_sign', 'carrier_sign', 'subject_to_count', 'missing_data']

    if 'export' in request.GET:
        excel_data = []
        print("EXCEL DATA", excel_data)
        for osd_group in osd_groups:
            row = [getattr(osd_group, field.name)
                   for field in fields if field.name not in excluded_fields]
            print(row, "ROW")
            excel_data.append(row)

        headers = [
            field.verbose_name for field in fields if field.name not in excluded_fields]

        workbook = Workbook()
        sheet = workbook.active

        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header

        for row_num, row in enumerate(excel_data, 2):
            for col_num, value in enumerate(row, 1):
                sheet.cell(row=row_num, column=col_num).value = value

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=osd_validation_data.xlsx'
        workbook.save(response)
        return response

    return render(request, 'osd_validation_engine.html', {'page_obj': page_obj, 'data': data, 'fields': fields, 'search_query': search_query,'sort_by': sort_by,
                                                            'filter_field': filter_field,'excluded_fields': excluded_fields,'date_fields': date_fields, 
                                                            'deduction_fields': deduction_fields,'backup_fields': backup_fields,'invoice_fields': invoice_fields,'POD_fields': POD_fields,
                                                            'calculated_fields': calculated_fields,'rca_fields': rca_fields, 'total_deductions': total_deductions,
                                                            'num_rows': num_rows, 'total_invalid': total_invalid,'invalid_rows': invalid_rows,
                                                            'total_valid': total_valid,'valid_rows': valid_rows,'rca_split': rca_split
                                                        })


@login_required(login_url="/login/")
def delete_osd_data(request):
    if request.method == 'POST':
        form = DeleteDataForm(request.POST)
        if form.is_valid():
            delete_deduction = form.cleaned_data['delete_deduction']
            delete_backup = form.cleaned_data['delete_backup']
            delete_invoice_data = form.cleaned_data['delete_invoice_data']
            delete_validation = form.cleaned_data['delete_validation']
            
            if delete_deduction:
                deduction_data.objects.all().delete()            
            if delete_backup:
                backup_data.objects.all().delete()            
            if delete_invoice_data:
                mv.invoice_data.objects.all().delete()            
            if delete_validation:
                validation.objects.all().delete()
    else:
        form = DeleteDataForm()
    
    context = {
        'form': form,
    }

    return render(request, 'delete_osd_data.html', context)        
            
            
    
def calculate_total_deductions(queryset):
# Sum of Deducted_Amount, handling None values
    total_deductions = queryset.aggregate(Sum('deducted_amount', output_field=models.DecimalField()))['deducted_amount__sum'] or Decimal(0)

    # Count the number of rows in the queryset
    num_rows = queryset.count()
    # Round off the result to two decimal places
    total_deductions = round(total_deductions, 2)

    return total_deductions , num_rows


def calculate_total_invalid(queryset):
    # Sum of INVALID AMOUNT
    total_invalid = queryset.aggregate(Sum('invalid_amount'))['invalid_amount__sum'] or 0
    
    # invalid_rows = queryset.count()
    invalid_rows = queryset.filter(validation_status="Invalid").count()
    # Round off the result to two decimal places
    total_invalid = round(total_invalid, 2)

    return total_invalid,invalid_rows


def calculate_total_valid(queryset):
    # Total Deductions - Total Invalid
    total_deductions, _ = calculate_total_deductions(queryset)
    total_invalid, _ = calculate_total_invalid(queryset)
    # Count the number of valid rows in the queryset
    valid_rows = queryset.filter(validation_status="Valid").count()

    # Round off the result to two decimal places
    total_valid = round(total_deductions - total_invalid, 2)

    return total_valid,valid_rows


def calculate_rca_split(queryset):
    # Initialize sums for the new detailed reasons
    deduction_is_taken_at_higher_price_sum = Decimal(0.0)
    pod_has_partial_shortages_sum =Decimal(0.0)
    pod_has_no_shortages_sum = Decimal(0.0)
    deducted_sku_is_not_invoiced_sum = Decimal(0.0)
    # price_change_not_communicated_sum = Decimal(0.0)

    # Iterate through the queryset and accumulate sums for each new reason
    for item in queryset:
        detailed_reason = item.invalid_reason
        invalid_amount = item.invalid_amount  
        if 'Deduction is taken at higher price' in detailed_reason:
            deduction_is_taken_at_higher_price_sum += invalid_amount
        elif 'POD has partial shortages' in detailed_reason:
            pod_has_partial_shortages_sum += invalid_amount 
        elif 'POD has no shortages' in detailed_reason:
            pod_has_no_shortages_sum += invalid_amount
        elif 'Deducted SKU is not invoiced' in detailed_reason:
            deducted_sku_is_not_invoiced_sum += invalid_amount    

    # Convert Decimal sums to float before constructing the dictionary
    sums = {
        'deduction_is_taken_at_higher_price': float( deduction_is_taken_at_higher_price_sum),
        'pod_has_partial_shortages': float(pod_has_partial_shortages_sum),
        'pod_has_no_shortages':float(pod_has_no_shortages_sum),
        'deducted_sku_is_not_invoiced': float(deducted_sku_is_not_invoiced_sum),
    }

    return sums


from django.db import transaction as db_transaction  # Rename the transaction import
@login_required(login_url="/login/")
def refresh_workflow(request):
    # Start a database transaction
    with transaction.atomic():
        # Delete all existing data from the Workflow table
        workflow.objects.all().delete()

        # Fetch new DeductionsData where priority is 'Yes' and not yet in Workflow
        new_deductions = deduction_data.objects.filter(priority='Yes').exclude(
            invoice_number__in=workflow.objects.values_list('invoice_number', flat=True)
        )

        # Prepare a list for bulk creation
        workflow_instances = []
        
        for deduction in new_deductions:
            # Aggregate validation data for the current invoice number
            validation_agg = validation.objects.filter(invoice_number=deduction.invoice_number).aggregate(
                sum_deducted_qty=Sum('deducted_qty'),
                billed_qty_sum=Sum('billed_qty'),
                avg_deducted_price_per_qty=Avg('deducted_price_per_qty'),
                avg_gross_price_per_qty=Avg('gross_price_per_qty'),
                avg_net_price_per_qty=Avg('net_price_per_qty'),
                max_carrier=Max('carrier'),
                sum_shortage=Sum('shortage'),
                sum_damage=Sum('damage'),
                sum_returns=Sum('returns'),
                sum_overage=Sum('overage'),
                sum_net_shortage=Sum('net_shortage'),
                max_customer_sign=Max('customer_sign'),
                max_carrier_sign=Max('carrier_sign'),
                max_subject_to_count=Max('subject_to_count'),
                max_no_shortage_in_pod=Max('no_shortage_in_pod'),
                max_partial_shortage_in_pod=Max('partial_shortage_in_pod'),
                max_deducted_sku_is_not_invoiced=Max('deducted_sku_is_not_invoiced'),
                max_pricing_variance=Max('pricing_variance'),
                max_validation_status=Max('validation_status'),
                sum_invalid_amount=Sum('invalid_amount'),
                sum_valid_amount=Sum('valid_amount'),
                max_invalid_reason=Max('invalid_reason'),
                max_product_substitution=Max('product_substitution'),
                max_combined_shipment=Max('combined_shipment'),
                max_order_split=Max('order_split'),
                max_load_sequencing=Max('load_sequencing'),
                max_unit_of_measurement=Max('unit_of_measurement'),
                max_deducted_at_higher_price=Max('deducted_at_higher_price'),
            )

            # Check if validation exists for the invoice number
            validation_pending = not any(validation_agg.values())

            # Create a new workflow instance
            workflow_instance = workflow(
                invoice_number=deduction.invoice_number,
                standard_customer=deduction.standard_customer,
                deduction_reference=deduction.deduction_reference,
                deducted_amount=deduction.deduction_amount,
                deducted_price_per_qty=validation_agg['avg_deducted_price_per_qty'] if not validation_pending else 0.0,
                deduction_date=deduction.deduction_date,
                deduction_reason=deduction.deduction_reason,
                deducted_qty=validation_agg['sum_deducted_qty'] if not validation_pending else 0,
                billed_qty=validation_agg['billed_qty_sum'] if not validation_pending else 0,
                gross_price_per_qty=validation_agg['avg_gross_price_per_qty'] if not validation_pending else 0.0,
                net_price_per_qty=validation_agg['avg_net_price_per_qty'] if not validation_pending else 0.0,
                carrier=validation_agg['max_carrier'] if not validation_pending else 'Validation Pending',
                shortage=validation_agg['sum_shortage'] if not validation_pending else 0.0,
                damage=validation_agg['sum_damage'] if not validation_pending else 0.0,
                returns=validation_agg['sum_returns'] if not validation_pending else 0.0,
                overage=validation_agg['sum_overage'] if not validation_pending else 0.0,
                net_shortage=validation_agg['sum_net_shortage'] if not validation_pending else 0.0,
                customer_sign=validation_agg['max_customer_sign'] if not validation_pending else 'Pending',
                carrier_sign=validation_agg['max_carrier_sign'] if not validation_pending else 'Pending',
                subject_to_count=validation_agg['max_subject_to_count'] if not validation_pending else 'Pending',
                no_shortage_in_pod=validation_agg['max_no_shortage_in_pod'] if not validation_pending else 'Pending',
                partial_shortage_in_pod=validation_agg['max_partial_shortage_in_pod'] if not validation_pending else 'Pending',
                deducted_sku_is_not_invoiced=validation_agg['max_deducted_sku_is_not_invoiced'] if not validation_pending else 'Pending',
                pricing_variance=validation_agg['max_pricing_variance'] if not validation_pending else 'Pending',
                validation_status=validation_agg['max_validation_status'] if not validation_pending else 'Pending',
                invalid_amount=validation_agg['sum_invalid_amount'] if not validation_pending else 0.0,
                valid_amount=validation_agg['sum_valid_amount'] if not validation_pending else 0.0,
                invalid_reason=validation_agg['max_invalid_reason'] if not validation_pending else 'Pending',
                billback_date=None,
                product_substitution=validation_agg['max_product_substitution'] if not validation_pending else 'Data not available',
                combined_shipment=validation_agg['max_combined_shipment'] if not validation_pending else 'Data not available',
                order_split=validation_agg['max_order_split'] if not validation_pending else 'Data not available',
                load_sequencing=validation_agg['max_load_sequencing'] if not validation_pending else 'Data not available',
                unit_of_measurement=validation_agg['max_unit_of_measurement'] if not validation_pending else 'Data not available',
                deducted_at_higher_price=validation_agg['max_deducted_at_higher_price'] if not validation_pending else 'Data not available',
            )

            workflow_instances.append(workflow_instance)

        # Perform a bulk create to save all instances at once
        workflow.objects.bulk_create(workflow_instances)
        update_missing_data()

        messages.success(request, "Workflow data has been refreshed successfully!")
        return redirect('workflow_view')  # Adjust the redirect URL to your workflow page

def update_missing_data():
    # Fetch all records where validation_status is 'Pending'
    pending_validations = workflow.objects.filter(validation_status="Pending")

    for workflow_item in pending_validations:
        # Check for missing data in invoice, pod, or backup
        if not mm.invoice_data.objects.filter(invoice_number=workflow_item.invoice_number).exists():
            workflow_item.invalid_reason = "Missing Invoice Data"
            workflow_item.validation_status = "Pending Validation"
        elif not pod_detail.objects.filter(invoice_number=workflow_item.invoice_number).exists():
            workflow_item.invalid_reason = "Missing POD Data"
            workflow_item.validation_status = "Pending Validation"
        elif not backup_data.objects.filter(invoice_number=workflow_item.invoice_number).exists():
            workflow_item.invalid_reason = "Missing Backup Data"
            workflow_item.validation_status = "Pending Validation"
        else:
            workflow_item.validation_status = "Pending Research"  # All data available
        
        # Save changes to the workflow item
        workflow_item.save()

    # Return success message or log it
    print("All 'Pending' validations updated successfully.")


@login_required(login_url="/login/")
def workflow_view(request):
    # Define the fields
    action_fields = ['actions']
    date_fields = ['deduction_date', 'billback_date']
    deduction_fields = ['standard_customer', 'deduction_reference', 'invoice_number', 'deduction_amount', 'deduction_date', 'deduction_reason','deducted_qty', 'deducted_price_per_qty' ]
    backup_fields = []
    invoice_fields = ['billed_qty', 'gross_price_per_qty', 'net_price_per_qty', 'carrier']
    POD_fields = ['shortage', 'damage', 'returns', 'overage', 'net_shortage','customer_sign', 'carrier_sign', 'subject_to_count' ]
    calculated_fields = ['no_shortage_in_pod', 'partial_shortage_in_pod', 
                         'deducted_sku_is_not_invoiced', 'pricing_variance', 
                         'validation_status', 'invalid_amount', 'valid_amount', 'invalid_reason','billback_package','billback_date','billback_id','customer_response', 'repayment_reference', 'repayment_amount', 'final_status']
    rca_fields = ['product_substitution', 'combined_shipment', 'order_split', 'load_sequencing', 'unit_of_measurement', 'deducted_at_higher_price']

    # Get the filter parameters from the query string
    search_query = request.GET.get('search', '')
    # Initialize sort_by variable
    sort_by = request.GET.get('sort', 'invoice_number')
    filter_field = request.GET.get('filter', '')  # Get the filter field value

    # Filter the validation objects based on the search query
    osd_groups = workflow.objects.filter(
        Q(invoice_number__icontains=search_query)  # Apply search query filter
    )

    # Pagination
    data = Paginator(osd_groups, 10)
    page = request.GET.get('page')
    page_obj = data.get_page(page)
    fields = workflow._meta.fields

    # Define excluded fields
    excluded_fields = ['id','ids','dedution_reference', 'invalid_amt_1',  'invalid_amt_2',
                          'invalid_amt_3',  'invalid_amt_4',  'missing_data']

    if 'export' in request.GET:
        excel_data = []
        print("EXCEL DATA", excel_data)
        for osd_group in osd_groups:
            row = [getattr(osd_group, field.name)
                   for field in fields if field.name not in excluded_fields]
            print(row, "ROW")
            excel_data.append(row)

        headers = [
            field.verbose_name for field in fields if field.name not in excluded_fields]

        workbook = Workbook()
        sheet = workbook.active

        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header

        for row_num, row in enumerate(excel_data, 2):
            for col_num, value in enumerate(row, 1):
                sheet.cell(row=row_num, column=col_num).value = value

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=osd_validation_data.xlsx'
        workbook.save(response)
        return response
    workflows = workflow.objects.all()
    for work in workflows:
        work.attachments = work.get_attachments()

    return render(request, 'workflow.html', {'page_obj': page_obj, 'data': data, 'fields': fields, 'search_query': search_query,'sort_by': sort_by,
                                                            'filter_field': filter_field,'excluded_fields': excluded_fields,'date_fields': date_fields, 
                                                            'deduction_fields': deduction_fields,'backup_fields': backup_fields,'invoice_fields': invoice_fields,'POD_fields': POD_fields,
                                                            'calculated_fields': calculated_fields,'rca_fields': rca_fields,'action_fields': action_fields, 'workflows': workflows,
                                                        })

from django.urls import reverse
from django.utils import timezone

from decimal import Decimal, InvalidOperation
from django.utils import timezone
from django.contrib import messages

def edit_workflow(request, workflow_id):
    workflow_item = get_object_or_404(workflow, id=workflow_id)

    if request.method == 'POST':
        # Update the fields with the data from the form
        for field in ['carrier', 'billback_id', 'customer_response', 'repayment_reference', 'final_status']:
            setattr(workflow_item, field, request.POST.get(field))

        # Handle billback_date separately to avoid format errors
        billback_date_str = request.POST.get('billback_date')
        if billback_date_str:  # If a date is provided
            try:
                # Parse the date string to a date object
                billback_date = timezone.datetime.strptime(billback_date_str, '%Y-%m-%d').date()
                workflow_item.billback_date = billback_date
            except ValueError:
                messages.error(request, "Invalid date format for billback date. Please use YYYY-MM-DD.")
                return redirect(reverse('edit_workflow', args=[workflow_id]))
        else:  # If no date is provided, set it to None
            workflow_item.billback_date = None

        # Handle decimal fields
        decimal_fields = ['repayment_amount']  # Add other decimal fields here
        for field in decimal_fields:
            value = request.POST.get(field)
            if value:  # If a value is provided
                try:
                    # Convert to Decimal
                    decimal_value = Decimal(value)
                    setattr(workflow_item, field, decimal_value)
                except (InvalidOperation, ValueError):
                    messages.error(request, f"Invalid value for {field}. Please enter a valid decimal number.")
                    return redirect(reverse('edit_workflow', args=[workflow_id]))
            else:  # If no value is provided, set it to None
                setattr(workflow_item, field, None)

        # Save the updated workflow item
        workflow_item.save()

    return redirect(reverse('workflow_view'))




def generate_billback_document(invoice_data):
    # Your implementation for generating the billback document
    # For example, create a simple text representation of the invoice data
    context = {
        'deduction_amount': invoice_data['deduction_amount'],
        'invalid_amount': invoice_data['invalid_amount'],
        'invoice_number': invoice_data['invoice_number'],
        'po_number': invoice_data['po_number'],
        'account': invoice_data['account'],
        'customer_name': invoice_data['customer_name']
    }
    # Use Django template system to render the billback document
    return render_to_string('billback_template.txt', context)

def create_billback_package(invoice_number):
    """
    Generate a billback package (ZIP file) for the given invoice number,
    including the billback template document and related files.
    """
    # Directory paths for pod, invoice, and backup
    pod_dir = os.path.join(settings.BASE_DIR, 'Uploads', 'pod')
    invoice_dir = os.path.join(settings.BASE_DIR, 'Uploads', 'invoice')
    backup_dir = os.path.join(settings.BASE_DIR, 'Uploads', 'backup')

    # Create ZIP archive in memory
    zip_buffer = io.BytesIO()
    zip_filename = f'billback_package_{invoice_number}.zip'

    with zipfile.ZipFile(zip_buffer, 'w') as zip_file:
        folder_name = f'{invoice_number}/'

        # Fetch all invoice instances related to this invoice_number
        invoice_instances = workflow.objects.filter(invoice_number=invoice_number, validation_status='Invalid')

        for invoice_instance in invoice_instances:
            # Fetch related PO instance for the first invoice (or handle as needed)
            po_instance = mm.invoice_data.objects.filter(invoice_number=invoice_number).first()
            deduction_instance = deduction_data.objects.filter(invoice_number=invoice_number).first()

            # Prepare invoice data for billback document
            invoice_data = {
                'deduction_amount': invoice_instance.deducted_amount,
                'invalid_amount': invoice_instance.invalid_amount,
                'invoice_number': invoice_instance.invoice_number,
                'po_number': po_instance.order_number if po_instance else 'Not Available',
                'account': deduction_instance.customer_account,
                'customer_name': deduction_instance.standard_customer
            }

            # Generate billback document and add it to the ZIP
            billback_content = generate_billback_document(invoice_data)
            zip_file.writestr(os.path.join(folder_name, f'billback_document_{invoice_instance.id}.txt'), billback_content)

            # Add related files (POD, Invoice, Backup) to the ZIP
            for directory in [pod_dir, invoice_dir, backup_dir]:
                for file in os.listdir(directory):
                    if invoice_number in file:
                        zip_file.write(os.path.join(directory, file), os.path.join(folder_name, file))

    zip_buffer.seek(0)

    # Save the ZIP file to the media directory
    billback_path = os.path.join(settings.MEDIA_ROOT, 'billback_packages', zip_filename)
    os.makedirs(os.path.dirname(billback_path), exist_ok=True)  # Ensure the directory exists
    with open(billback_path, 'wb') as f:
        f.write(zip_buffer.getvalue())

    # Return the relative URL for saving in the database or displaying as a link
    return os.path.join(settings.MEDIA_URL, 'billback_packages', zip_filename)

from django.contrib import messages
from django.shortcuts import redirect

def create_billback_package_view(request):
    # Fetch deductions that have a validation status of 'Invalid'
    invalid_deductions = workflow.objects.filter(validation_status='Invalid')

    # Use a set to track unique invoice numbers
    processed_invoices = set()

    # Check if there are any invalid deductions
    if invalid_deductions.exists():
        for deduction in invalid_deductions:
            # Check if the invoice number has already been processed
            if deduction.invoice_number not in processed_invoices:
                # Generate a billback package for each unique invalid deduction
                billback_package_url = create_billback_package(deduction.invoice_number)
                
                # Update the deduction with the generated package URL
                deduction.billback_package = billback_package_url
                deduction.save()  # Save the instance to the database

                # Add the invoice number to the set to prevent duplicates
                processed_invoices.add(deduction.invoice_number)

        messages.success(request, "Billback packages have been created successfully!")
    else:
        messages.info(request, "No invalid deductions found to create billback packages.")

    return redirect('workflow_view')  # Adjust this to your workflow page URL


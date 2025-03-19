from main.utils import *
from main import models as mm
from django.db import connection
from django.db import IntegrityError
from openpyxl import Workbook
import time
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.http import JsonResponse, HttpResponseServerError
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import condition
from decimal import Decimal
from django.core.exceptions import ValidationError
from django.db.models import Q
from django.db import transaction
from django.db.models.functions import Concat
from django.db.models import Sum, F, Value
from django.db.models import Count, OuterRef, Subquery
from datetime import datetime
from .models import *
from django.shortcuts import render, HttpResponse, redirect
from django.contrib import messages
import pandas as pd
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.template import loader
from dateutil import parser
from pathlib import Path
from django.core.paginator import Paginator
from django.shortcuts import render
import logging
logger = logging.getLogger(__name__)
from .forms import *
from osd import models as co


@login_required(login_url="/login/")
def deductions_upload(request, excel_file, context):

    try:
        if excel_file:
            # Load the Excel data into a pandas DataFrame
            df = pd.read_csv(excel_file)
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df.shape[0]
            instances = []

            # Iterate over each row in the DataFrame and validate the data
            for _, row in df.iterrows():
                try:

                    ids = row['id']
                    # deduction_date = row['deduction_date'] if 'deduction_date' in row else None
                    deduction_date = row.get('deduction_date', None)
                    if pd.isnull(deduction_date):
                        deduction_date = None
                    else:
                        deduction_date = parse_date(deduction_date)
                    customer_name = row['customer_name'] if 'customer_name' in row else ''
                    customer_account = row['customer_account'] if 'customer_account' in row else ''
                    standard_customer = row['standard_customer'] if 'standard_customer' in row else ''
                    deduction_reference = row['deduction_reference'] if 'deduction_reference' in row else ''
                    invoice_number = row['invoice_number']
                    deduction_amount = row['deduction_amount']
                    if deduction_amount is None or deduction_amount == '':
                        deduction_amount = 0
                    else:
                        deduction_amount = convert_to_decimal_V2(
                            row['deduction_amount'])
                    deduction_reason = row['deduction_reason'] if 'deduction_reason' in row else ''
                    payment_number = row['payment_number']
                    backup_status = row['backup_status'] if 'backup_status' in row else ''
                    invoice_status = row['invoice_status'] if 'invoice_status' in row else ''
                    validation = row['validation'] if 'validation' in row else ''                    
                    invalid_amount = clean_value(row['invalid_amount'], 0.0)
                    valid_amount = clean_value(row['valid_amount'], 0.0)
                    billback = row['billback'] if 'billback' in row else ''
                    billback_date = row['billback_date'] if row['billback_date'] not in [None, ''] else None 
                    if pd.isnull(billback_date):
                        billback_date = None
                    else:
                        billback_date = parse_date(billback_date)
                                      
                    billback_amount = clean_value(row['billback_amount'], 0.0)
                    recovery_status = row['recover_status'] if 'recover_status' in row else ''                     
                    recovered_amount = row.get('recovered_amount')                    
                    if pd.isna(recovered_amount):
                        recovered_amount = 0.0
                    else:
                        recovered_amount = clean_value_v3(recovered_amount, 0.0)
                     

                    # Check if the record already exists in the model
                    if deduction_data.objects.filter(invoice_number=invoice_number, payment_number=payment_number, deduction_amount=deduction_amount).exists():
                        continue

                    # Create an instance of the model
                    instance = deduction_data(
                        ids=ids,
                        deduction_date=deduction_date,
                        customer_name=customer_name,
                        customer_account=customer_account,
                        standard_customer=standard_customer,
                        deduction_reference=deduction_reference,
                        invoice_number=invoice_number,
                        deduction_amount=deduction_amount,
                        deduction_reason=deduction_reason,
                        payment_number=payment_number,
                        backup_status=backup_status,
                        invoice_status=invoice_status,
                        validation=validation,                        
                        invalid_amount=invalid_amount,
                        valid_amount=valid_amount,
                        billback=billback,
                        billback_date=billback_date,
                        billback_amount=billback_amount,
                        recovery_status=recovery_status,
                        recovered_amount=recovered_amount
                    )
                    instances.append(instance)

                except ValueError as e:
                    print(f"ValueError for row {row}: {e}")
                except Exception as e:
                    print(f"Unexpected error for row {row}: {e}")

            # Bulk insert the instances
            deduction_data.objects.bulk_create(instances)
            end_time = timezone.now()
            total_time = end_time - start_time

            # Update upload status
            current_status = 'completed'
            request.session['upload_status'] = 'success'

            # Log the upload details
            uploaded_csv = mm.UploadedCSV.objects.create(
                user=request.user,
                file=excel_file,
                num_rows=df.shape[0],
                start_time=start_time,
                end_time=end_time,
                total_time=total_time,
                current_status=current_status
            )
            return redirect('backup_upload')

        else:
            context['error'] = 'No Excel file was uploaded.'

    except pd.errors.EmptyDataError:
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)


@login_required(login_url="/login/")
def deductions_view(request):

    date_field = []
    excluded_fields = ['id', 'invalid_amount', 'valid_amount','validation','billback','billback_date','billback_amount',
                       'recovered_amount','payment_number','backup_status','invoice_status','recovery_status']
    # Setup pagination
    data = Paginator(deduction_data.objects.all(), 10)
    page = request.GET.get('page')
    page_obj = data.get_page(page)
    fields = deduction_data._meta.fields
    field_names = [
        field.name for field in fields if field.name not in excluded_fields]
    print(field_names)

    return render(request, 'pricing_deductions_view.html', {'date_field': date_field, 'page_obj': page_obj, 'data': data, 'field_names': field_names})


@login_required(login_url="/login/")
def backup_upload(request, excel_file, context):
    try:
        if excel_file:
            # Load the Excel data into a pandas DataFrame
            df = pd.read_csv(excel_file, dtype={'invoice_number': str,'sku': str})
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df.shape[0]
            instances = []

            # Iterate over each row in the DataFrame and validate the data
            for _, row in df.iterrows():
                try:

                    ids = row['id']
                    standard_customer = row['standard_customer'] if 'standard_customer' in row else ''
                    invoice_number = row['invoice_number'] if 'invoice_number' in row else ''
                    sku = row['sku'] if 'sku' in row else ''
                    deducted_amount = row['deducted_amount']
                    if deducted_amount is None or deducted_amount == '':
                        deducted_amount = 0.0
                    else:
                        deducted_amount = convert_to_decimal_V2(
                            deducted_amount)
                        if deducted_amount is None or deducted_amount != deducted_amount:
                            deducted_amount = 0.0
                    deducted_qty = row['deducted_qty']
                    if deducted_qty is None or deducted_qty == '':
                        deducted_qty = 0
                    else:
                        deducted_qty = convert_to_decimal_V2(deducted_qty)
                        if deducted_qty is None or deducted_qty != deducted_qty:
                            deducted_qty = 0
                    deducted_price_per_qty = row['deducted_price_per_qty']
                    if deducted_price_per_qty is None or deducted_price_per_qty == '':
                        deducted_price_per_qty = 0.0
                    else:
                        deducted_price_per_qty = convert_to_decimal_V2(
                            deducted_price_per_qty)
                        if deducted_price_per_qty is None or deducted_price_per_qty != deducted_price_per_qty:
                            deducted_price_per_qty = 0.0

                    deduction_date = row['deduction_date'] if 'deduction_date' in row else None
                    if pd.isnull(deduction_date):
                        deduction_date = None
                    else:
                        deduction_date = parse_date(row['deduction_date'])

                    invoice_location = row['invoice_location'] if 'invoice_location' in row else ''
                    reason_code = row['reason_code'] if 'reason_code' in row else ''
                    sub_reason_code = row['sub_reason_code'] if 'sub_reason_code' in row else ''
                    deduction_reason = row['deduction_reason'] if 'deduction_reason' in row else ''
                    backup_status = row['backup_status'] if 'backup_status' in row else ''

                    # Check if the record already exists in the model
                    if backup_data.objects.filter(invoice_number=invoice_number, sku=sku).exists():
                        continue

                    # Create an instance of the model
                    instance = backup_data(
                        ids=ids,
                        standard_customer=standard_customer,
                        invoice_number=invoice_number,
                        sku=sku,
                        deducted_amount=deducted_amount,
                        deducted_qty=deducted_qty,
                        deducted_price_per_qty=deducted_price_per_qty,
                        deduction_date=deduction_date,
                        invoice_location=invoice_location,
                        reason_code=reason_code,
                        sub_reason_code=sub_reason_code,
                        deduction_reason=deduction_reason,
                        backup_status=backup_status
                    )
                    instances.append(instance)

                except ValueError as e:
                    print(f"ValueError for row {row}: {e}")
                except Exception as e:
                    print(f"Unexpected error for row {row}: {e}")

            # Bulk insert the instances
            # backup_data.objects.bulk_create(instances)
            if instances:
                backup_data.objects.bulk_create(instances)
            end_time = timezone.now()
            total_time = end_time - start_time

            # Update upload status
            current_status = 'completed'
            request.session['upload_status'] = 'success'

            # Log the upload details
            uploaded_csv = mm.UploadedCSV.objects.create(
                user=request.user,
                file=excel_file,
                num_rows=df.shape[0],
                start_time=start_time,
                end_time=end_time,
                total_time=total_time,
                current_status=current_status
            )
            return redirect('backup_upload')

        else:
            context['error'] = 'No Excel file was uploaded.'

    except pd.errors.EmptyDataError:
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)


@login_required(login_url="/login/")
def backup_view(request):
    date_field = ['deduction_date']
    excluded_fields = ['id', 'invoice_location', 'sub_reason_code','backup_status', 'reason_code']
    # Setup pagination
    data = Paginator(backup_data.objects.all(), 10)
    page = request.GET.get('page')
    page_obj = data.get_page(page)
    fields = backup_data._meta.fields
    field_names = [
        field.name for field in fields if field.name not in excluded_fields]
    print(field_names)

    return render(request, 'pricing_backup.html', {'date_field': date_field, 'page_obj': page_obj, 'data': data, 'field_names': field_names})

@login_required(login_url="/login/")
def price_change_upload(request, excel_file, context):

    try:
        if excel_file:
            # Load the Excel data into a pandas DataFrame
            df = pd.read_csv(excel_file, dtype={'sku': str})
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df.shape[0]
            instances = []
            # Iterate over each row in the DataFrame and validate the data
            for _, row in df.iterrows():
                sku = row['sku']
                unit_price = row['unit_price']
                if pd.isnull(unit_price):                    
                    unit_price = None  
                pack_price = row['pack_price']
                if pd.isnull(pack_price):                    
                    pack_price = None  
                case_price = row['case_price']
                if pd.isnull(case_price):                    
                    case_price = None  
                units_per_case = row['units_per_case']                              
                units_per_pack = row['units_per_pack']
                packs_per_case = row['packs_per_case'] 
                
                effective_date = row['effective_date']                
                if pd.notnull(effective_date):
                    try:
                        effective_date = datetime.strptime(effective_date, '%m/%d/%Y').date()
                    except ValueError:
                        print(f"Error parsing effective_date: {effective_date}")
                        effective_date = None
                else:
                    effective_date = None
                
                communication_date = row['communication_date']
                if pd.notnull(communication_date):
                    try:
                        communication_date = datetime.strptime(communication_date, '%m/%d/%Y').date()
                    except ValueError:
                        print(f"Error parsing communication_date: {communication_date}")
                        communication_date = None
                else:
                    communication_date = None
                
                buyer_approved = row['buyer_approved']

                # Check if the record already exists in the price_change model
                if price_change.objects.filter(sku=sku, units_per_case=units_per_case).exists():
                    # Skip duplicates
                    continue
                instance = price_change(
                    sku=sku,
                    unit_price=unit_price,
                    pack_price=pack_price,
                    case_price=case_price,
                    units_per_case=units_per_case,
                    units_per_pack=units_per_pack,
                    packs_per_case=packs_per_case,
                    effective_date=effective_date,
                    communication_date=communication_date,
                    buyer_approved=buyer_approved,
                )
                instances.append(instance)

            # Bulk insert the instances
            price_change.objects.bulk_create(instances)
            end_time = timezone.now()
            total_time = end_time - start_time

            # Update status to 'completed'
            current_status = 'completed'
            request.session['upload_status'] = 'success'

            #update upload status
            uploaded_csv = mm.UploadedCSV.objects.create(
                user=request.user,
                file=excel_file,
                num_rows=df.shape[0],
                start_time=start_time,
                end_time=end_time,
                total_time=total_time,
                current_status=current_status
            )
            return redirect('backup_upload')
        else:
            # Handle the case when excel_file is not present
            print("File not present")
            context['error'] = 'No file was uploaded.'

    except KeyError as e:
        # Handle missing columns
        print(f"Missing column: {e}")
        context['error'] = f'Missing column: {e}'

    except pd.errors.EmptyDataError:
        print("EmptyDataError caught")
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        print(f'ParserError caught: {e}')
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        print(f'ValidationError caught: {e}')        
        context['error'] = f'Validation error: {e}'
        print(f'Row data: {row}')
    
    
    except Exception as e:
        # Handle other unexpected exceptions
        print(f'Unexpected error caught: {e}')
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)

def price_change_data(request):
    
    #setup pagination
    data = Paginator(price_change.objects.all(),20)
    page = request.GET.get('page')
    page_obj = data.get_page(page)
    fields = price_change._meta.fields

    return render(request, 'price_change.html', {'page_obj': page_obj, 'data': data, 'fields': fields})  

@login_required(login_url="/login/")
def transform_pricing_data(request):

    if request.method == 'POST':
        form = TransformDataForm(request.POST)
        if form.is_valid():
            backup_data_calculation = form.cleaned_data['backup_data_calculation']
            group_pricing = form.cleaned_data['group_pricing'] 
            validation_data_calculation = form.cleaned_data['validation_data_calculation'] 

            if backup_data_calculation:
                backup_data_calculate(request)
                return redirect('transform_pricing_data') 

            if group_pricing:
                combine_pricing_data(request) 
                # test.delay()
                return redirect('transform_pricing_data')     

            if validation_data_calculation:
                validation_data_calculate(request)                 
                return redirect('transform_pricing_data')             
    else:
        form = TransformDataForm()
    
    context = {
        'form': form,
    }
    return render(request, 'transform_pricing_data.html', context)

def backup_data_calculate(request):
    try:
        print("performing execution")        
        # SQL query for updating inv_sku by concatenating deduction_reference and sku
        update_inv_sku_sql = """
            UPDATE pricing_backup_data
            SET inv_sku = invoice_number || sku
            WHERE invoice_number IS NOT NULL AND sku IS NOT NULL;
        """

        # Execute the raw SQL queries sequentially
        with connection.cursor() as cursor:            
            cursor.execute(update_inv_sku_sql)

        # Save the instance to persist the changes in the database
        transaction.commit()
        return redirect('transform_pricing_data')
    except Exception as e:
        print(f'An unexpected error occurred: {e}')
        error_message = f'An unexpected error occurred: {e}'
        return HttpResponseServerError(error_message)
    
def combine_pricing_data(request):

    try:
        print("Attempting to combine pricing data...")
        with connection.cursor() as cursor:
            sql_query = """
                INSERT INTO pricing_pricing_validation (
                    inv_sku, ids, invoice_number, sku, deducted_amount, deducted_qty, deducted_price_per_qty, deduction_reason,
                    ids, deduction_date, standard_customer, deduction_reference, payment_number,
                    invoice_date, billed_qty, gross_price, net_price, gross_price_per_qty, net_price_per_qty,
                    unit_price, pack_price, case_price, units_per_case, units_per_pack, packs_per_case, effective_date, communication_date, buyer_approved,
                    customer_expected_price, validation_status, invalid_amount, detailed_reason, customer_expected_price_lower_than_billed,
                    price_change_not_communicated, price_change_communicated_late, unit_of_measurement_issue
                )
                SELECT
                    COALESCE(bd.inv_sku, ''),
                    COALESCE(bd.ids, ''),
                    COALESCE(bd.invoice_number, ''),
                    COALESCE(bd.sku, ''),
                    COALESCE(bd.deducted_amount, 0),
                    COALESCE(bd.deducted_qty, 0),
                    COALESCE(bd.deducted_price_per_qty, 0),
                    COALESCE(bd.deduction_reason, ''),
                    COALESCE(dd.ids, ''),
                    COALESCE(dd.deduction_date, ''),
                    COALESCE(dd.standard_customer, ''),
                    COALESCE(dd.deduction_reference, ''),
                    COALESCE(dd.payment_number, ''),
                    COALESCE(inv.invoice_date, ''),
                    COALESCE(inv.billed_qty, 0),
                    COALESCE(inv.gross_price, 0),
                    COALESCE(inv.net_price, 0),
                    COALESCE(inv.gross_price_per_qty, 0),
                    COALESCE(inv.net_price_per_qty, 0),
                    COALESCE(pc.unit_price, 0),
                    COALESCE(pc.pack_price, 0),
                    COALESCE(pc.case_price, 0),
                    COALESCE(pc.units_per_case, 0),
                    COALESCE(pc.units_per_pack, 0),
                    COALESCE(pc.packs_per_case, 0),
                    COALESCE(pc.effective_date, ''),
                    COALESCE(pc.communication_date, ''),
                    COALESCE(pc.buyer_approved, 'no'),
                    NULL, -- Placeholder for customer_expected_price,
                    '', -- Placeholder for validation_status,
                    NULL, -- Placeholder for invalid_amount,
                    '', -- Placeholder for detailed_reason, 
                    '', -- Placeholder for customer_expected_price_lower_than_billed,               
                    '', -- Placeholder for price_change_not_communicated,
                    '', -- Placeholder for price_change_communicated_late,
                    '' -- Placeholder for unit_of_measurement_issue 

                    
                FROM
                    pricing_backup_data bd
                    LEFT JOIN pricing_deduction_data dd ON bd.ids = dd.ids
                    LEFT JOIN main_invoice_data inv ON bd.inv_sku = inv.inv_sku
                    LEFT JOIN pricing_price_change pc ON bd.sku = pc.sku
            """
            print("Executing SQL query...")
            cursor.execute(sql_query)
    except IntegrityError as e:
        error_message = f"Integrity Error: {e}"
        print(error_message)
        return render(request, 'transform_pricing_data.html', {'error_message': error_message})
    except Exception as e:
        print(f"Error: {e}")
        error_message = f"Database Error: {e}"
        return render(request, 'transform_pricing_data.html', {'error_message': error_message})
    
    print("combine_pricing_data executed successfully.")
    return render(request, 'transform_pricing_data.html', {'success_message': 'Pricing data combined successfully.'})

@login_required(login_url="/login/")
def pricing_validation_view(request):
    try:        
        date_fields = ['deduction_date', 'invoice_date', 'effective_date', 'communication_date']
        backup_data_fields = ['invoice_number', 'deducted_qty', 'deducted_price_per_qty','deduction_reason','sku', 'deducted_amount'  ]
        deduction_data_fields = [ 'deduction_reference','standard_customer', 'deduction_date', 'deduction_reference','payment_number']
        invoice_data_fields = ['gross_price','gross_price_per_qty', 'net_price_per_qty']
        price_change_fields = ['unit_price', 'pack_price', 'case_price',  'effective_date', 'communication_date']
        calculated_fields = ['customer_expected_price','customer_expected_price_lower_than_billed', 'price_change_not_communicated', 'price_change_communicated_late', 'unit_of_measurement_issue','Qty_ded_vs_Billed_issue','mod_issue','validation_status', 'invalid_amount','valid_amount', 'detailed_reason']
        
        # Get the filter parameters from the query string
        search_query = request.GET.get('search', '')
        sort_by = request.GET.get('sort', 'inv_sku')
        filter_field = request.GET.get('filter', ' ')  
        
        print("Search Query:", search_query)
        print("Sort By:", sort_by)
        print("Filter Field:", filter_field)

        
        
        # Filter the PromoGroup objects based on the search query and filter field
        pricing_groups = pricing_validation.objects.filter(
            Q(inv_sku__icontains=search_query) |  # Apply search query filter
            Q(validation_status__icontains=search_query)    # Apply filter field filter
        )

        
        
        # Sort the PromoGroup objects based on the sort_by parameter
        pricing_groups = pricing_groups.order_by(sort_by)
        
        
        
        # Calculate the required values for the summary
        total_deductions = calculate_total_deductions(pricing_groups)
        total_deductions, num_rows = calculate_total_deductions(pricing_groups)
        total_invalid, invalid_rows= calculate_total_invalid(pricing_groups)
        total_valid,valid_rows = calculate_total_valid(pricing_groups)
        rca_split = calculate_rca_split(pricing_groups) 
        print(total_deductions,total_invalid,total_valid,"SUMMARY")    

        
        # Pagination
        data = Paginator(pricing_groups, 10)
        page = request.GET.get('page')
        page_obj = data.get_page(page)
        fields = pricing_validation._meta.fields

        # # Define excluded fields
        excluded_fields = ['inv_sku','id','ids', 'payment_ref_no', 'po_number', 'invoice_date', 'invoice_location',
                        'item_description', 'sub_reason_code', 'dispute_id', 'customer_number', 'company_code',
                        'deduction_amount', 'product_category', 'payment_number', 'billed_qty', 'billed_price', 'oi_deals',
                        'other_allowance', 'cash_discount', 'billed_others', 'net_price', 'units_per_case',
                        'units_per_pack', 'packs_per_case', 'buyer_approved', 'qty_ded_vs_Billed_issue','mod_issue', 'customer_expected_price_lower_than_billed']

        if 'export' in request.GET:
            excel_data = []
            for pricing_groups in pricing_groups:
                row = [getattr(pricing_groups, field.name) for field in fields if field.name not in excluded_fields]
                excel_data.append(row)

            headers = [field.verbose_name for field in fields if field.name not in excluded_fields]

            # Create a new workbook and get the active sheet
            workbook = Workbook()
            sheet = workbook.active

            # Write headers to the first row
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num).value = header

            # Write data to the remaining rows
            for row_num, row in enumerate(excel_data, 2):
                for col_num, value in enumerate(row, 1):
                    sheet.cell(row=row_num, column=col_num).value = value

            # Set the appropriate response headers
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=pricing_validation_data.xlsx'

            # Save the workbook to the response
            workbook.save(response)

            return response


        return render(request, 'pricing_validation_engine.html', {'page_obj': page_obj, 'data': data, 'fields': fields,
                                                        'search_query': search_query, 'sort_by': sort_by,'excluded_fields': excluded_fields,
                                                        'filter_field': filter_field,  'date_fields': date_fields, 'backup_data_fields':backup_data_fields,
                                                        'deduction_data_fields': deduction_data_fields, 'invoice_data_fields': invoice_data_fields, 'price_change_fields':price_change_fields,
                                                        'calculated_fields':calculated_fields,'total_deductions': total_deductions,'num_rows': num_rows,
                                                        'total_invalid': total_invalid,'invalid_rows':invalid_rows,'total_valid': total_valid,'valid_rows':valid_rows,'rca_split': rca_split                                                       
                                                            })
    
   
    except Exception as e:
        # Handle other unexpected exceptions
        print(f'An unexpected error occurred: {e}')
        error_message = f'An unexpected error occurred: {e}'
        return HttpResponseServerError(error_message)


def calculate_total_deductions(queryset):
# Sum of Deducted_Amount, handling None values
    total_deductions = queryset.aggregate(Sum('deducted_amount', output_field=models.DecimalField()))['deducted_amount__sum'] or Decimal(0)

    # Count the number of rows in the queryset
    num_rows = queryset.count()
    # Round off the result to two decimal places
    total_deductions = round(total_deductions, 2)

    return total_deductions , num_rows


def calculate_total_invalid(queryset):
    # Sum of INVALID AMOUNT
    total_invalid = queryset.aggregate(Sum('invalid_amount'))['invalid_amount__sum'] or 0
    
    # invalid_rows = queryset.count()
    invalid_rows = queryset.filter(validation_status="Invalid").count()
    # Round off the result to two decimal places
    total_invalid = round(total_invalid, 2)

    return total_invalid,invalid_rows


def calculate_total_valid(queryset):
    # Total Deductions - Total Invalid
    total_deductions, _ = calculate_total_deductions(queryset)
    total_invalid, _ = calculate_total_invalid(queryset)
    # Count the number of valid rows in the queryset
    valid_rows = queryset.filter(validation_status="Valid").count()

    # Round off the result to two decimal places
    total_valid = round(total_deductions - total_invalid, 2)

    return total_valid,valid_rows


def calculate_rca_split(queryset):
    # Initialize sums for the new detailed reasons
    unit_of_measurement_issue_sum = Decimal(0.0)
    price_change_communicated_on_time_sum =Decimal(0.0)
    price_change_communicated_late_sum = Decimal(0.0)
    price_change_not_communicated_sum = Decimal(0.0)
    # price_change_not_communicated_sum = Decimal(0.0)

    # Iterate through the queryset and accumulate sums for each new reason
    for item in queryset:
        detailed_reason = item.detailed_reason  
        invalid_amount = item.invalid_amount  
        # valid_amount = item.valid_amount
        if 'Unit of Measurement Issue' in detailed_reason:
            unit_of_measurement_issue_sum += invalid_amount
        elif 'Price change was communicated on time' in detailed_reason:
            # price_change_communicated_late_sum += invalid_amount
            price_change_communicated_on_time_sum += invalid_amount 
        elif 'Price change was communicated late' in detailed_reason:
            price_change_communicated_late_sum += invalid_amount
        elif 'Price change was not communicated' in detailed_reason:
            price_change_not_communicated_sum += invalid_amount    

    # Convert Decimal sums to float before constructing the dictionary
    sums = {
        'unit_of_measurement_issue': float(unit_of_measurement_issue_sum),
        'price_change_communicated_late': float(price_change_communicated_late_sum),
        'price_change_communicated_on_time':float(price_change_communicated_on_time_sum),
        'price_change_not_communicated': float(price_change_not_communicated_sum),
    }

    return sums

from main import models as mm
@login_required(login_url="/login/")
def delete_data(request):
    if request.method == 'POST':
        form = DeleteDataForm(request.POST)
        if form.is_valid():
            
            delete_deductions = form.cleaned_data['delete_deductions']
            delete_invoice_data = form.cleaned_data['delete_invoice_data']
            delete_backup = form.cleaned_data['delete_backup']
            delete_price_change = form.cleaned_data['delete_price_change']           
            delete_pricing_validation = form.cleaned_data['delete_pricing_validation']
            
            if delete_deductions:
                deduction_data.objects.all().delete()
                
            if delete_invoice_data:
                mm.invoice_data.objects.all().delete()
                
            if delete_backup:
                backup_data.objects.all().delete()
            
            if delete_price_change:
                price_change.objects.all().delete()
            
            if delete_pricing_validation:
                pricing_validation.objects.all().delete()

           

    else:
        form = DeleteDataForm()
    
    context = {
        'form': form,
    }

    return render(request, 'delete_pricing_data.html', context)


def validation_data_calculate(request):
    try:
        print("performing execution")

        update_customer_expected_price_sql = """
            UPDATE pricing_pricing_validation
            SET customer_expected_price = net_price_per_qty - deducted_price_per_qty;
        """
        
        update_customer_expected_price_lower_than_billed_sql = """
            UPDATE pricing_pricing_validation
            SET customer_expected_price_lower_than_billed = CASE
                WHEN customer_expected_price < net_price_per_qty
                THEN TRUE
                ELSE FALSE
            END;
        """

        update_price_change_not_communicated_sql = """
            UPDATE pricing_pricing_validation
            SET price_change_not_communicated = CASE
                WHEN communication_date IS NULL OR communication_date = ''
                THEN 'Valid'
                ELSE 'Invalid'
            END;
        """

        update_price_change_communicated_late_sql = """
            UPDATE pricing_pricing_validation
            SET price_change_communicated_late = CASE
                WHEN communication_date IS NULL OR communication_date = ''
                THEN 'Valid'
                ELSE 
                    CASE
                        WHEN DATE(communication_date, '+60 day') < invoice_date
                        THEN 'Valid'
                        ELSE 'Invalid'
                    END
            END;
            """

        update_unit_of_measurement_issue_sql = """
            UPDATE pricing_pricing_validation
            SET unit_of_measurement_issue = CASE
                WHEN qty_ded_vs_Billed_issue = 'TRUE' OR mod_issue = 0
                THEN 'Invalid'
                ELSE 'Valid'
            END;
        """

        update_qty_ded_vs_billed_issue_sql = """
            UPDATE pricing_pricing_validation
            SET qty_ded_vs_Billed_issue = CASE  
                WHEN billed_qty > 1 AND deducted_qty = 1
                THEN 'TRUE'
                ELSE 'FALSE'
            END;
        """

        update_mod_issue_sql = """
            UPDATE pricing_pricing_validation
            SET mod_issue = ROUND(MOD(deducted_price_per_qty / customer_expected_price, 1), 2);
        """
        
        update_validation_status_sql = """
            UPDATE pricing_pricing_validation
            SET validation_status = 
                CASE
                    WHEN unit_of_measurement_issue = 'Invalid'
                    THEN 'Invalid'
                    WHEN price_change_communicated_late = 'Invalid'
                    THEN 'Invalid'
                    WHEN price_change_not_communicated = 'Valid'
                    THEN 'Valid'
                    WHEN price_change_communicated_late = 'Valid'
                    THEN 'Valid'
                    ELSE validation_status
                END
            WHERE validation_status IS NULL OR validation_status = '';
        """
        
        update_invalid_amount_sql = """
            UPDATE pricing_pricing_validation
            SET invalid_amount = CASE
                WHEN validation_status = 'Invalid'
                THEN deducted_amount
                ELSE 0
            END;
        """
        
        update_valid_amount_sql = """
            UPDATE pricing_pricing_validation
            SET valid_amount = deducted_amount - invalid_amount;
        """
        
        update_detailed_reason_sql = """
            UPDATE pricing_pricing_validation
            SET detailed_reason = CASE
                WHEN unit_of_measurement_issue = 'Invalid' THEN 'Unit of Measurement Issue'
                WHEN price_change_communicated_late = 'Invalid' THEN 'Price change was communicated on time'
                WHEN price_change_not_communicated = 'Valid' THEN 'Price change was not communicated'
                WHEN price_change_not_communicated = 'Invalid' AND price_change_communicated_late = 'Valid' THEN 'Price change was communicated late'
                ELSE 'Valid Deduction'
            END;
        """

        # Execute the raw SQL queries sequentially
        with connection.cursor() as cursor:
            cursor.execute(update_customer_expected_price_sql)
            cursor.execute(update_customer_expected_price_lower_than_billed_sql)
            cursor.execute(update_price_change_not_communicated_sql)
            cursor.execute(update_price_change_communicated_late_sql)            
            cursor.execute(update_qty_ded_vs_billed_issue_sql)
            cursor.execute(update_mod_issue_sql)
            cursor.execute(update_unit_of_measurement_issue_sql)
            cursor.execute(update_validation_status_sql)
            cursor.execute(update_invalid_amount_sql)
            cursor.execute(update_valid_amount_sql)
            cursor.execute(update_detailed_reason_sql)

        # Save the instance to persist the changes in the database
        transaction.commit()
        print("No exception occurred. Redirecting to 'transform_pricing_data'.")
        return redirect('transform_pricing_data')

    except Exception as e:
        print(f'An unexpected error occurred: {e}')
        error_message = f'An unexpected error occurred: {e}'
        return HttpResponseServerError(error_message)
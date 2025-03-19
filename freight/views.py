from main.utils import *
from main import models as mm
from django.db import connection
from django.db import IntegrityError
from openpyxl import Workbook
import time
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import condition
from decimal import Decimal
from django.core.exceptions import ValidationError
from django.db.models import Q
from django.db import transaction
from django.db.models.functions import Concat
from django.db.models import Sum, F, Value
from django.db.models import Count, OuterRef, Subquery
from datetime import datetime
from .models import *
# from .forms import *
from django.shortcuts import render, HttpResponse, redirect
from django.contrib import messages
import pandas as pd
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.template import loader
from dateutil import parser
from pathlib import Path
from django.core.paginator import Paginator
from django.shortcuts import render
import logging
logger = logging.getLogger(__name__)
from .forms import *
from osd import models as co

@login_required(login_url="/login/")
def deductions_upload(request, excel_file, context):
    try:
        if excel_file:
            # Load the Excel data into a pandas DataFrame
            df = pd.read_csv(excel_file)
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df.shape[0]
            instances = []

            # Iterate over each row in the DataFrame and validate the data
            for _, row in df.iterrows():
                try:

                    ids = row['id']
                    # deduction_date = row['deduction_date'] if 'deduction_date' in row else None
                    deduction_date = row.get('deduction_date', None)
                    if pd.isnull(deduction_date):
                        deduction_date = None
                    else:
                        deduction_date = parse_date(deduction_date)
                    customer_name = row['customer_name'] if 'customer_name' in row else ''
                    customer_account = row['customer_account'] if 'customer_account' in row else ''
                    standard_customer = row['standard_customer'] if 'standard_customer' in row else ''
                    deduction_reference = row['deduction_reference'] if 'deduction_reference' in row else ''
                    invoice_number = row['invoice_number']
                    deduction_amount = row['deduction_amount']
                    if deduction_amount is None or deduction_amount == '':
                        deduction_amount = 0
                    else:
                        deduction_amount = convert_to_decimal_V2(
                            row['deduction_amount'])
                    deduction_reason = row['deduction_reason'] if 'deduction_reason' in row else ''
                    payment_number = row['payment_number']
                    backup_status = row['backup_status'] if 'backup_status' in row else ''
                    invoice_status = row['invoice_status'] if 'invoice_status' in row else ''
                    validation = row['validation'] if 'validation' in row else ''                    
                    invalid_amount = clean_value(row['invalid_amount'], 0.0)
                    valid_amount = clean_value(row['valid_amount'], 0.0)
                    billback = row['billback'] if 'billback' in row else ''
                    billback_date = row['billback_date'] if row['billback_date'] not in [None, ''] else None 
                    if pd.isnull(billback_date):
                        billback_date = None
                    else:
                        billback_date = parse_date(billback_date)
                                      
                    billback_amount = clean_value(row['billback_amount'], 0.0)
                    recovery_status = row['recover_status'] if 'recover_status' in row else ''                     
                    recovered_amount = row.get('recovered_amount')                    
                    if pd.isna(recovered_amount):
                        recovered_amount = 0.0
                    else:
                        recovered_amount = clean_value_v3(recovered_amount, 0.0)
                     

                    # Check if the record already exists in the model
                    if deduction_data.objects.filter(invoice_number=invoice_number, payment_number=payment_number, deduction_amount=deduction_amount).exists():
                        continue

                    # Create an instance of the model
                    instance = deduction_data(
                        ids=ids,
                        deduction_date=deduction_date,
                        customer_name=customer_name,
                        customer_account=customer_account,
                        standard_customer=standard_customer,
                        deduction_reference=deduction_reference,
                        invoice_number=invoice_number,
                        deduction_amount=deduction_amount,
                        deduction_reason=deduction_reason,
                        payment_number=payment_number,
                        backup_status=backup_status,
                        invoice_status=invoice_status,
                        validation=validation,                        
                        invalid_amount=invalid_amount,
                        valid_amount=valid_amount,
                        billback=billback,
                        billback_date=billback_date,
                        billback_amount=billback_amount,
                        recovery_status=recovery_status,
                        recovered_amount=recovered_amount
                    )
                    instances.append(instance)

                except ValueError as e:
                    print(f"ValueError for row {row}: {e}")
                except Exception as e:
                    print(f"Unexpected error for row {row}: {e}")

            # Bulk insert the instances
            deduction_data.objects.bulk_create(instances)
            end_time = timezone.now()
            total_time = end_time - start_time

            # Update upload status
            current_status = 'completed'
            request.session['upload_status'] = 'success'

            # Log the upload details
            uploaded_csv = mm.UploadedCSV.objects.create(
                user=request.user,
                file=excel_file,
                num_rows=df.shape[0],
                start_time=start_time,
                end_time=end_time,
                total_time=total_time,
                current_status=current_status
            )
            return redirect('backup_upload')

        else:
            context['error'] = 'No Excel file was uploaded.'

    except pd.errors.EmptyDataError:
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)

@login_required(login_url="/login/")
def deductions_view(request):
    date_field = []
    excluded_fields = ['id', 'invalid_amount', 'valid_amount','validation','billback','billback_date','billback_amount',
                       'recovered_amount','payment_number','backup_status','invoice_status','recovery_status']
    # Setup pagination
    data = Paginator(deduction_data.objects.all(), 10)
    page = request.GET.get('page')
    page_obj = data.get_page(page)
    fields = deduction_data._meta.fields
    field_names = [
        field.name for field in fields if field.name not in excluded_fields]
    print(field_names)

    return render(request, 'freight_deductions_view.html', {'date_field': date_field, 'page_obj': page_obj, 'data': data, 'field_names': field_names})


@login_required(login_url="/login/")
def backup_upload(request, excel_file, context):
    try:
        if excel_file:
            # Load the Excel data into a pandas DataFrame
            df = pd.read_csv(excel_file)
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df.shape[0]
            instances = []

            # Iterate over each row in the DataFrame and validate the data
            for _, row in df.iterrows():
                try:

                    ids = row['id']
                    standard_customer = row['standard_customer'] if 'standard_customer' in row else ''
                    invoice_number = row['invoice_number'] if 'invoice_number' in row else ''
                    sku = row['sku'] if 'sku' in row else ''
                    deducted_amount = row['deducted_amount']
                    if deducted_amount is None or deducted_amount == '':
                        deducted_amount = 0.0
                    else:
                        deducted_amount = convert_to_decimal_V2(
                            deducted_amount)
                        if deducted_amount is None or deducted_amount != deducted_amount:
                            deducted_amount = 0.0
                    deducted_qty = row['deducted_qty']
                    if deducted_qty is None or deducted_qty == '':
                        deducted_qty = 0
                    else:
                        deducted_qty = convert_to_decimal_V2(deducted_qty)
                        if deducted_qty is None or deducted_qty != deducted_qty:
                            deducted_qty = 0
                    deducted_price_per_qty = row['deducted_price_per_qty']
                    if deducted_price_per_qty is None or deducted_price_per_qty == '':
                        deducted_price_per_qty = 0.0
                    else:
                        deducted_price_per_qty = convert_to_decimal_V2(
                            deducted_price_per_qty)
                        if deducted_price_per_qty is None or deducted_price_per_qty != deducted_price_per_qty:
                            deducted_price_per_qty = 0.0

                    deduction_date = row['deduction_date'] if 'deduction_date' in row else None
                    if pd.isnull(deduction_date):
                        deduction_date = None
                    else:
                        deduction_date = parse_date(row['deduction_date'])

                    invoice_location = row['invoice_location'] if 'invoice_location' in row else ''
                    reason_code = row['reason_code'] if 'reason_code' in row else ''
                    sub_reason_code = row['sub_reason_code'] if 'sub_reason_code' in row else ''
                    deduction_reason = row['deduction_reason'] if 'deduction_reason' in row else ''
                    backup_status = row['backup_status'] if 'backup_status' in row else ''

                    # Check if the record already exists in the model
                    if backup_data.objects.filter(invoice_number=invoice_number, sku=sku).exists():
                        continue

                    # Create an instance of the model
                    instance = backup_data(
                        ids=ids,
                        standard_customer=standard_customer,
                        invoice_number=invoice_number,
                        sku=sku,
                        deducted_amount=deducted_amount,
                        deducted_qty=deducted_qty,
                        deducted_price_per_qty=deducted_price_per_qty,
                        deduction_date=deduction_date,
                        invoice_location=invoice_location,
                        reason_code=reason_code,
                        sub_reason_code=sub_reason_code,
                        deduction_reason=deduction_reason,
                        backup_status=backup_status
                    )
                    instances.append(instance)

                except ValueError as e:
                    print(f"ValueError for row {row}: {e}")
                except Exception as e:
                    print(f"Unexpected error for row {row}: {e}")

            # Bulk insert the instances
            # backup_data.objects.bulk_create(instances)
            if instances:
                backup_data.objects.bulk_create(instances)
            end_time = timezone.now()
            total_time = end_time - start_time

            # Update upload status
            current_status = 'completed'
            request.session['upload_status'] = 'success'

            # Log the upload details
            uploaded_csv = mm.UploadedCSV.objects.create(
                user=request.user,
                file=excel_file,
                num_rows=df.shape[0],
                start_time=start_time,
                end_time=end_time,
                total_time=total_time,
                current_status=current_status
            )
            return redirect('backup_upload')

        else:
            context['error'] = 'No Excel file was uploaded.'

    except pd.errors.EmptyDataError:
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)


@login_required(login_url="/login/")
def backup_view(request):
    date_field = ['deduction_date']
    excluded_fields = ['id', 'invoice_location', 'sub_reason_code','sku', 'deducted_qty', 'deducted_price_per_qty','backup_status', 'reason_code']
    # Setup pagination
    data = Paginator(backup_data.objects.all(), 10)
    page = request.GET.get('page')
    page_obj = data.get_page(page)
    fields = backup_data._meta.fields
    field_names = [
        field.name for field in fields if field.name not in excluded_fields]
    print(field_names)

    return render(request, 'freight_backup.html', {'date_field': date_field, 'page_obj': page_obj, 'data': data, 'field_names': field_names})


@login_required(login_url="/login/")
def invoice_upload(request, excel_file=None, context={}):

    try:
        if excel_file:
            # Load the Excel data into a pandas DataFrame
            df = pd.read_csv(excel_file, dtype={'invoice_number': str,'sku': str})
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df.shape[0]
            instances = []

            # Define a function to convert NaN to default values
            def get_value(value, default=0):
                return default if pd.isna(value) else value

            # Iterate over each row in the DataFrame and validate the data
            for _, row in df.iterrows():
                try:
                    # Extract fields as per the model
                    inv_sku = row['concat'] if 'concat' in row else ''
                    invoice_number = row['invoice_number'] if 'invoice_number' in row else ''
                    sku = row['sku'] if 'sku' in row else ''
                    billed_qty = clean_value(row['billed_qty'], 0)
                    gross_price = row['gross_price']
                    if gross_price is None or gross_price == '':
                        gross_price = 0.0
                    else:
                        gross_price = convert_to_decimal_V2(gross_price)
                        if gross_price is None or gross_price != gross_price:
                            gross_price = 0.0
                    oi_deal = clean_value(row['oi_deal'], 0.0)
                    promo_allowance = clean_value(row['promo_allowance'], 0.0)
                    cash_discount = clean_value(row['cash_discount'], 0.0)
                    others = clean_value(row['others'], 0.0)
                    net_price = clean_value(row['net_price'], 0.0)
                    gross_price_per_qty = clean_value(
                        row['gross_price_per_qty'], 0.0)
                    net_price_per_qty = clean_value(
                        row['net_price_per_qty'], 0.0)
                    order_number = row['order_number'] if 'order_number' in row else ''
                    bol = row['bol'] if 'bol' in row else ''
                    carrier = row['carrier'] if 'carrier' in row else ''

                    # Check if the record already exists in the model
                    if invoice_data.objects.filter(inv_sku=inv_sku, order_number=order_number).exists():
                        continue

                    # Create an instance of the model
                    instance = invoice_data(
                        inv_sku=inv_sku,
                        invoice_number=invoice_number,
                        sku=sku,
                        billed_qty=billed_qty,
                        gross_price=gross_price,
                        oi_deal=oi_deal,
                        promo_allowance=promo_allowance,
                        cash_discount=cash_discount,
                        others=others,
                        net_price=net_price,
                        gross_price_per_qty=gross_price_per_qty,
                        net_price_per_qty=net_price_per_qty,
                        order_number=order_number,
                        bol=bol,
                        carrier=carrier
                    )
                    instances.append(instance)

                except ValueError as e:
                    print(f"ValueError for row {row}: {e}")
                except Exception as e:
                    print(f"Unexpected error for row {row}: {e}")

            # Bulk insert the instances
            if instances:
                invoice_data.objects.bulk_create(instances)

            end_time = timezone.now()
            total_time = end_time - start_time

            # Update upload status
            current_status = 'completed'
            request.session['upload_status'] = 'success'

            # Log the upload details
            uploaded_csv = mm.UploadedCSV.objects.create(
                user=request.user,
                file=excel_file,
                num_rows=df.shape[0],
                start_time=start_time,
                end_time=end_time,
                total_time=total_time,
                current_status=current_status
            )
            return redirect('backup_upload')

        else:
            context['error'] = 'No Excel file was uploaded.'

    except pd.errors.EmptyDataError:
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)

@login_required(login_url="/login/")
def invoice_view(request):
    date_field = []
    excluded_fields = ['id', 'inv_sku']
    # Setup pagination
    data = Paginator(invoice_data.objects.all(), 10)
    page = request.GET.get('page')
    page_obj = data.get_page(page)
    fields = invoice_data._meta.fields
    field_names = [
        field.name for field in fields if field.name not in excluded_fields]
    print(field_names)

    return render(request, 'invoice_data.html', {'date_field': date_field, 'page_obj': page_obj, 'data': data, 'field_names': field_names})

@login_required(login_url="/login/")
def edi_master_upload(request, excel_file, context):
    try:
        if excel_file:
            # Load the Excel data into a pandas DataFrame
            df = pd.read_csv(excel_file)
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df.shape[0]
            instances = []

            # Iterate over each row in the DataFrame and validate the data
            for _, row in df.iterrows():
                try:

                    # deduction_date = row['deduction_date'] if 'deduction_date' in row else None
                    deduction_reason = row.get('deduction_reason', None)
                    freight_code_master = row.get('freight_code_master', None)
                    
                     

                    # Check if the record already exists in the model
                    if edi_master.objects.filter(deduction_reason=deduction_reason, freight_code_master=freight_code_master).exists():
                        continue

                    # Create an instance of the model
                    instance = edi_master(
                        deduction_reason=deduction_reason,
                        freight_code_master=freight_code_master
                        
                    )
                    instances.append(instance)

                except ValueError as e:
                    print(f"ValueError for row {row}: {e}")
                except Exception as e:
                    print(f"Unexpected error for row {row}: {e}")

            # Bulk insert the instances
            edi_master.objects.bulk_create(instances)
            end_time = timezone.now()
            total_time = end_time - start_time

            # Update upload status
            current_status = 'completed'
            request.session['upload_status'] = 'success'

            # Log the upload details
            uploaded_csv = mm.UploadedCSV.objects.create(
                user=request.user,
                file=excel_file,
                num_rows=df.shape[0],
                start_time=start_time,
                end_time=end_time,
                total_time=total_time,
                current_status=current_status
            )
            return redirect('backup_upload')

        else:
            context['error'] = 'No Excel file was uploaded.'

    except pd.errors.EmptyDataError:
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)


@login_required(login_url="/login/")
def edi_actual_upload(request, excel_file, context):
    try:
        if excel_file:
            # Load the Excel data into a pandas DataFrame
            df = pd.read_csv(excel_file)
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df.shape[0]
            instances = []

            # Iterate over each row in the DataFrame and validate the data
            for _, row in df.iterrows():
                try:

                    invoice_number = row.get('invoice_number', None)
                    deduction_reason = row.get('deduction_reason', None)
                    freight_code_actual = row.get('freight_code_actual', None)
                    
                     

                    # Check if the record already exists in the model
                    if edi_actual.objects.filter(invoice_number=invoice_number, deduction_reason=deduction_reason, freight_code_actual=freight_code_actual).exists():
                        continue

                    # Create an instance of the model
                    instance = edi_actual(
                        invoice_number=invoice_number,
                        deduction_reason=deduction_reason,
                        freight_code_actual=freight_code_actual
                        
                    )
                    instances.append(instance)

                except ValueError as e:
                    print(f"ValueError for row {row}: {e}")
                except Exception as e:
                    print(f"Unexpected error for row {row}: {e}")

            # Bulk insert the instances
            edi_actual.objects.bulk_create(instances)
            end_time = timezone.now()
            total_time = end_time - start_time

            # Update upload status
            current_status = 'completed'
            request.session['upload_status'] = 'success'

            # Log the upload details
            uploaded_csv = mm.UploadedCSV.objects.create(
                user=request.user,
                file=excel_file,
                num_rows=df.shape[0],
                start_time=start_time,
                end_time=end_time,
                total_time=total_time,
                current_status=current_status
            )
            return redirect('backup_upload')

        else:
            context['error'] = 'No Excel file was uploaded.'

    except pd.errors.EmptyDataError:
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)


@login_required(login_url="/login/")
def freight_communication_upload(request, excel_file, context):
    try:
        if excel_file:
            # Load the Excel data into a pandas DataFrame
            df = pd.read_csv(excel_file)
            start_time = timezone.now()

            # Set initial status
            current_status = 'started'
            total_rows = df.shape[0]
            instances = []

            # Iterate over each row in the DataFrame and validate the data
            for _, row in df.iterrows():
                try:
                    
                    customer_account = row.get('customer_account', None)
                    standard_customer = row.get('standard_customer', None)
                    communicated_lane = row.get('lane', None)

                    communicated_rate = row['rate']
                    if communicated_rate is None or communicated_rate == '':
                        communicated_rate = 0.0
                    else:
                        communicated_rate = convert_to_decimal_V2(communicated_rate)
                        if communicated_rate is None or communicated_rate != communicated_rate:
                            communicated_rate = 0.0

                    communicated_weight = row['weight']
                    if communicated_weight is None or communicated_weight == '':
                        communicated_weight = 0.0
                    else:
                        communicated_weight = convert_to_decimal_V2(communicated_weight)
                        if communicated_weight is None or communicated_weight != communicated_weight:
                            communicated_weight = 0.0
                    
                     

                    # Check if the record already exists in the model
                    if freight_communication.objects.filter(customer_account=customer_account,standard_customer=standard_customer,communicated_lane=communicated_lane).exists():
                        continue

                    # Create an instance of the model
                    instance = freight_communication(
                        customer_account=customer_account,
                        standard_customer=standard_customer,
                        communicated_lane=communicated_lane,
                        communicated_rate=communicated_rate,
                        communicated_weight=communicated_weight

                        
                    )
                    instances.append(instance)

                except ValueError as e:
                    print(f"ValueError for row {row}: {e}")
                except Exception as e:
                    print(f"Unexpected error for row {row}: {e}")

            # Bulk insert the instances
            freight_communication.objects.bulk_create(instances)
            end_time = timezone.now()
            total_time = end_time - start_time

            # Update upload status
            current_status = 'completed'
            request.session['upload_status'] = 'success'

            # Log the upload details
            uploaded_csv = mm.UploadedCSV.objects.create(
                user=request.user,
                file=excel_file,
                num_rows=df.shape[0],
                start_time=start_time,
                end_time=end_time,
                total_time=total_time,
                current_status=current_status
            )
            return redirect('backup_upload')

        else:
            context['error'] = 'No Excel file was uploaded.'

    except pd.errors.EmptyDataError:
        context['error'] = 'The uploaded file is empty.'

    except pd.errors.ParserError as e:
        context['error'] = f'Error parsing the CSV file: {e}'

    except ValidationError as e:
        context['error'] = f'Validation error: {e}'

    except Exception as e:
        context['error'] = f'An unexpected error occurred: {e}'

    return render(request, 'backup_upload.html', context)

def transform_freight_data(request):
    if request.method == 'POST':
        form = TransformDataForm(request.POST)
        if form.is_valid():
            freight_validation = form.cleaned_data['freight_validation']
            freight_calculations = form.cleaned_data['freight_calculations']
            if freight_validation:
                combine_freight_data(request)
                return redirect('transform_freight_data')
            if freight_calculations:
                update_calculated_fields(request)
                return redirect('transform_freight_data')
    else:
        form = TransformDataForm()
    context = {
        'form': form,
    }
    print("Rendering transform_freight_data.html...")
    return render(request, 'transform_freight_data.html', context)

def combine_freight_data(request):
    try:
        print("Attempting to combine Freight data...")
        with connection.cursor() as cursor:
            sql_query = """
                INSERT INTO freight_validation (
                    ids_bd, ids_dd, standard_customer, customer_account, deduction_reference, deduction_date,
                    invoice_number, deducted_amount, deduction_reason, billed_qty, freight_amount,
                    fuel_allowance, lane, freight_rate, total_freight, gross_weight, freight_code_master, freight_code_actual,
                    communicated_lane, communicated_rate, communicated_weight
                )
                SELECT DISTINCT
                    COALESCE(bd."ids", '') AS ids_bd,
                    COALESCE(dd."ids", '') AS ids_dd,
                    COALESCE(dd."standard_customer", '') AS standard_customer,
                    COALESCE(dd."customer_account", '') AS customer_account,
                    COALESCE(dd."deduction_reference", '') AS deduction_reference,
                    COALESCE(dd."deduction_date", '') AS deduction_date,
                    COALESCE(bd."invoice_number", '') AS invoice_number,
                    COALESCE(bd."deducted_amount", 0) AS deducted_amount,
                    COALESCE(bd."deduction_reason", '') AS deduction_reason,
                    COALESCE(id."billed_qty", 0) AS billed_qty,
                    COALESCE(id."freight_amount", 0) AS freight_amount,
                    COALESCE(id."fuel_allowance", 0) AS fuel_allowance,
                    COALESCE(id."lane", '') AS lane,
                    COALESCE(id."freight_rate", 0) AS freight_rate,
                    COALESCE(id."total_freight", 0) AS total_freight,
                    COALESCE(id."gross_weight", 0) AS gross_weight,
                    COALESCE(em."freight_code_master", '') AS freight_code_master,
                    COALESCE(ea."freight_code_actual", '') AS freight_code_actual,
                    COALESCE(fc."communicated_lane", '') AS communicated_lane,
                    COALESCE(fc."communicated_rate", '') AS communicated_rate,
                    COALESCE(fc."communicated_weight", '') AS communicated_weight
                FROM freight_backup_data bd
                LEFT JOIN freight_deduction_data dd 
                    ON dd."invoice_number" = bd."invoice_number"
                    AND dd."ids" = bd."ids"
                INNER JOIN (
                    SELECT 
                        "invoice_number",
                        SUM("billed_qty") AS billed_qty,
                        SUM("freight_amount") AS freight_amount,
                        SUM("fuel_allowance") AS fuel_allowance,
                        MAX("lane") AS lane,
                        MAX("freight_rate") AS freight_rate,
                        MAX("total_freight") AS total_freight,
                        MAX("gross_weight") AS gross_weight
                    FROM main_invoice_data
                    GROUP BY "invoice_number"
                ) id ON bd."invoice_number" = id."invoice_number"
                LEFT JOIN freight_edi_master em 
                    ON bd."deduction_reason" = em."deduction_reason"
                LEFT JOIN freight_edi_actual ea 
                    ON bd."deduction_reason" = ea."deduction_reason"
                LEFT JOIN freight_freight_communication fc 
                    ON dd."customer_account" = fc."customer_account"
                WHERE NOT EXISTS (
                    SELECT 1
                    FROM freight_validation val
                    WHERE val."invoice_number" = bd."invoice_number"
                    AND val."ids_bd" = bd."ids"
                );

            """
            print("Executing SQL query...")
            cursor.execute(sql_query)
    except IntegrityError as e:
        error_message = f"Integrity Error: {e}"
        print(error_message)
        return render(request, 'transform_freight_data.html', {'error_message': error_message})
    except Exception as e:
        print(f"Error: {e}")
        error_message = f"Database Error: {e}"
        return render(request, 'transform_freight_data.html', {'error_message': error_message})

    print("Freight data executed successfully.")
    return render(request, 'transform_osd_data.html', {'success_message': 'Freight data combined successfully.'})


def update_calculated_fields(self):
    # SQL queries to update calculated fields
    update_load_wise = """
            UPDATE freight_validation
            SET load_wise = CASE
                WHEN total_freight <> 0 THEN 'Invalid'
                ELSE 'Valid'
            END
            WHERE load_wise IS NULL;
        """

    update_invalid_amt_1 = """
            UPDATE freight_validation
            SET invalid_amt_1 = CASE
                WHEN load_wise = 'Invalid' THEN deducted_amount
                ELSE 0
            END
            WHERE invalid_amt_1 IS NULL;
        """

    update_invoice_count = """
        UPDATE freight_validation
        SET invoice_count = (
            SELECT COUNT(*)
            FROM freight_validation fv2
            WHERE fv2."invoice_number" = freight_validation."invoice_number"
        )
        WHERE EXISTS (
            SELECT 1
            FROM freight_validation fv2
            WHERE fv2."invoice_number" = freight_validation."invoice_number"
        );
    """
    
    update_duplicate = """
            UPDATE freight_validation
            SET duplicate = CASE
                WHEN invoice_count > 1 AND load_wise = 'Valid' THEN 'Invalid'
                ELSE 'Valid'
            END
            WHERE duplicate IS NULL;
        """

    update_invalid_amt_2 = """
            UPDATE freight_validation
            SET invalid_amt_2 = CASE
                WHEN duplicate = 'Invalid' THEN deducted_amount
                ELSE 0
            END
            WHERE invalid_amt_2 IS NULL;
        """

    update_rate_wise = """
            UPDATE freight_validation
            SET rate_wise = CASE
                WHEN freight_rate > communicated_rate AND duplicate = 'Valid' THEN 'Invalid'
                ELSE 'Valid'
            END
            WHERE rate_wise IS NULL;
        """

    update_invalid_amt_3 = """
            UPDATE freight_validation
            SET invalid_amt_3 = CASE
                WHEN rate_wise = 'Invalid' THEN deducted_amount
                ELSE 0
            END
            WHERE invalid_amt_3 IS NULL;
        """

    update_edi_mismatch = """
            UPDATE freight_validation
            SET edi_mismatch = CASE
                WHEN freight_code_master <> freight_code_actual AND rate_wise = 'Valid' THEN 'Invalid'
                ELSE 'Valid'
            END
            WHERE edi_mismatch IS NULL;
        """

    update_invalid_amt_4 = """
            UPDATE freight_validation
            SET invalid_amt_4 = CASE
                WHEN edi_mismatch = 'Invalid' THEN deducted_amount
                ELSE 0
            END
            WHERE invalid_amt_4 IS NULL;
        """
    update_validation_status = """
            UPDATE freight_validation
            SET validation_status = 
                CASE
                    WHEN load_wise = 'Invalid' THEN 'Invalid'
                    WHEN duplicate = 'Invalid' THEN 'Invalid'
                    WHEN rate_wise = 'Invalid' THEN 'Invalid'
                    WHEN edi_mismatch = 'Invalid' THEN 'Invalid'
                    ELSE 'Valid'
                END
            WHERE validation_status IS NULL OR validation_status = '';

        """
    update_final_rca = """
                UPDATE freight_validation
                SET final_rca =
                    CASE
                        WHEN final_rca IS NULL OR final_rca = ''
                        THEN
                            CASE
                                WHEN load_wise = 'Invalid'
                                THEN 'Freight is already given on the invoice'
                                WHEN duplicate = 'Invalid'
                                THEN 'Duplicate Freight Deduction'
                                WHEN rate_wise = 'Invalid'
                                THEN 'Freight is taken at a higher rate than communicated'
                                WHEN edi_mismatch = 'Invalid'
                                THEN 'EDI Mismatch'
                                ELSE 'Valid'
                            END
                        ELSE final_rca
                    END;
        """
    update_invalid_amt = """
                UPDATE freight_validation
                SET invalid_amt = CASE
                    WHEN invalid_amt_1 IS NOT NULL THEN COALESCE(invalid_amt_1, 0) + COALESCE(invalid_amt_2, 0) + COALESCE(invalid_amt_3, 0) + COALESCE(invalid_amt_4, 0)
                    ELSE 0
                END
                WHERE invalid_amt IS NULL OR invalid_amt = '';

        """

    update_valid_amt = """
                UPDATE freight_validation
                SET valid_amt = COALESCE(deducted_amount, 0) - COALESCE(invalid_amt, 0)
                WHERE valid_amt IS NULL OR valid_amt = '';

        """

    with connection.cursor() as cursor:
        cursor.execute(update_load_wise)
        cursor.execute(update_invalid_amt_1)
        cursor.execute(update_invoice_count)
        cursor.execute(update_duplicate)
        cursor.execute(update_invalid_amt_2)
        cursor.execute(update_rate_wise)
        cursor.execute(update_invalid_amt_3)
        cursor.execute(update_edi_mismatch)
        cursor.execute(update_invalid_amt_4)
        cursor.execute(update_validation_status)
        cursor.execute(update_final_rca)
        cursor.execute(update_invalid_amt)
        cursor.execute(update_valid_amt)

    return redirect('transform_freight_data')

@login_required(login_url="/login/")
def freight_validation_view(request):
    # Define the fields
    date_fields = ['deduction_date']
    deduction_fields = ['ids_dd', 'standard_customer','customer_account', 'deduction_reference','deduction_date']
    backup_fields = ['invoice_number','deducted_amount', 'deducted_qty', 'deducted_price_per_qty', 'deduction_reason']
    invoice_fields = ['billed_qty', 'freight_rate', 'freight_amount','total_freight', 'gross_weight']
    edi_fields = ['freight_code_master', 'freight_code_actual']
    freight_communication = ['communicated_rate', 'communicated_weight']
    calculated_fields = ['load_wise', 'invalid_amt_1', 'invoice_count', 'duplicate', 'invalid_amt_2',
                         'rate_wise', 'invalid_amt_3', 'edi_mismatch', 'invalid_amt_4',
                         'validation_status', 'invalid_amt', 'valid_amt', 'final_rca']

    # Get the filter parameters from the query string
    search_query = request.GET.get('search', '')
    # Initialize sort_by variable
    sort_by = request.GET.get('sort', 'invoice_number')
    filter_field = request.GET.get('filter', '')  # Get the filter field value

    # Filter the validation objects based on the search query
    freight_groups = validation.objects.filter(
        Q(invoice_number__icontains=search_query)  # Apply search query filter
    )

    # Sort the validation objects based on the sort_by parameter
    try:
        freight_groups = freight_groups.order_by(sort_by)
    except TypeError as e:
        print("Error during sorting:", e)

    # Calculate the required values for the summary
    total_deductions, num_rows = calculate_total_deductions(freight_groups)
    total_invalid, invalid_rows = calculate_total_invalid(freight_groups)
    total_valid, valid_rows = calculate_total_valid(freight_groups)
    rca_split = calculate_rca_split(freight_groups)

    # Pagination
    data = Paginator(freight_groups, 10)
    page = request.GET.get('page')
    page_obj = data.get_page(page)
    fields = validation._meta.fields

    # Define excluded fields
    excluded_fields = ['id','deduction_reference','deducted_qty','deducted_price_per_qty','gross_price_per_qty','net_price_per_qty','fuel_allowance',
                       'lane','communicated_lane','ids_dd','ids_bd']

    if 'export' in request.GET:
        excel_data = []
        print("EXCEL DATA", excel_data)
        for freight_group in freight_groups:
            row = [getattr(freight_group, field.name)
                   for field in fields if field.name not in excluded_fields]
            print(row, "ROW")
            excel_data.append(row)

        headers = [
            field.verbose_name for field in fields if field.name not in excluded_fields]

        workbook = Workbook()
        sheet = workbook.active

        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header

        for row_num, row in enumerate(excel_data, 2):
            for col_num, value in enumerate(row, 1):
                sheet.cell(row=row_num, column=col_num).value = value

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=osd_validation_data.xlsx'
        workbook.save(response)
        return response

    return render(request, 'freight_validation_view.html', {'page_obj': page_obj, 'data': data, 'fields': fields, 'search_query': search_query,'sort_by': sort_by,
                                                            'filter_field': filter_field,'excluded_fields': excluded_fields,'date_fields': date_fields, 
                                                            'deduction_fields': deduction_fields,'backup_fields': backup_fields,'invoice_fields': invoice_fields,'edi_fields':edi_fields,'freight_communication':freight_communication,
                                                            'calculated_fields': calculated_fields, 'total_deductions': total_deductions,
                                                            'num_rows': num_rows, 'total_invalid': total_invalid,'invalid_rows': invalid_rows,
                                                            'total_valid': total_valid,'valid_rows': valid_rows,'rca_split': rca_split
                                                        })


@login_required(login_url="/login/")
def delete_freight_data(request):
    if request.method == 'POST':
        form = DeleteDataForm(request.POST)
        if form.is_valid():
            delete_deduction = form.cleaned_data['delete_deduction']
            delete_backup = form.cleaned_data['delete_backup']
            delete_invoice_data = form.cleaned_data['delete_invoice_data']
            delete_validation = form.cleaned_data['delete_validation']
            
            if delete_deduction:
                deduction_data.objects.all().delete()            
            if delete_backup:
                backup_data.objects.all().delete()            
            if delete_invoice_data:
                co.invoice_data.objects.all().delete()          
            if delete_validation:
                validation.objects.all().delete()
    else:
        form = DeleteDataForm()
    
    context = {
        'form': form,
    }

    return render(request, 'delete_freight_data.html', context)        